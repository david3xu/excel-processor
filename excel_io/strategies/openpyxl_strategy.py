import os
import logging
from typing import Any, Dict, Iterator, List, Optional, Tuple

import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell import Cell
from openpyxl.utils import get_column_letter
from datetime import datetime, date

from ..interfaces import ExcelReaderInterface, SheetAccessorInterface, CellValueExtractorInterface
from .base_strategy import ExcelAccessStrategy

logger = logging.getLogger(__name__)


class OpenpyxlStrategy(ExcelAccessStrategy):
    """Strategy implementation using openpyxl for Excel access."""
    
    def create_reader(self, file_path: str, **kwargs) -> 'OpenpyxlReader':
        """
        Create an OpenpyxlReader for the specified Excel file.
        
        Args:
            file_path: Path to the Excel file
            **kwargs: Additional parameters for openpyxl
            
        Returns:
            OpenpyxlReader instance
            
        Raises:
            FileNotFoundError: If the file does not exist
            UnsupportedFileError: If the file cannot be handled by openpyxl
            ExcelAccessError: For other Excel access errors
        """
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"File not found: {file_path}")
        
        try:
            # Validate the file can be opened with openpyxl
            if not self.can_handle_file(file_path):
                raise UnsupportedFileError(f"File cannot be handled by openpyxl: {file_path}")
            
            read_only = kwargs.get('read_only', False)
            data_only = kwargs.get('data_only', True)
            
            return OpenpyxlReader(file_path, read_only=read_only, data_only=data_only)
        except Exception as e:
            if isinstance(e, FileNotFoundError) or isinstance(e, UnsupportedFileError):
                raise
            
            logger.error(f"Error creating openpyxl reader: {str(e)}")
            raise ExcelAccessError(f"Error creating openpyxl reader: {str(e)}") from e
    
    def can_handle_file(self, file_path: str) -> bool:
        """
        Determine if openpyxl can handle the specified file.
        
        Args:
            file_path: Path to the Excel file
            
        Returns:
            True if openpyxl can handle the file, False otherwise
        """
        if not os.path.exists(file_path):
            return False
        
        # Check file extension
        if not file_path.lower().endswith(('.xlsx', '.xlsm', '.xltx', '.xltm')):
            return False
        
        # Try to open the file with openpyxl to verify compatibility
        try:
            # Just open the workbook to check if it's supported
            # Use read_only=True for efficiency
            wb = openpyxl.load_workbook(file_path, read_only=True)
            wb.close()
            return True
        except Exception as e:
            logger.debug(f"Openpyxl cannot handle file {file_path}: {str(e)}")
            return False
    
    def get_strategy_name(self) -> str:
        """
        Get the name of this strategy.
        
        Returns:
            'openpyxl'
        """
        return 'openpyxl'
    
    def get_strategy_capabilities(self) -> Dict[str, bool]:
        """
        Get the capabilities supported by this strategy.
        
        Returns:
            Dictionary of capabilities
        """
        return {
            'merged_cells': True,
            'formulas': True,
            'styles': True,
            'complex_structures': True,
            'large_files': False  # Not ideal for very large files
        }


class OpenpyxlReader(ExcelReaderInterface):
    """Excel reader implementation using openpyxl."""
    
    def __init__(self, file_path: str, read_only: bool = False, data_only: bool = True):
        """
        Initialize the openpyxl reader.
        
        Args:
            file_path: Path to the Excel file
            read_only: Whether to open the file in read-only mode
            data_only: Whether to read values instead of formulas
        """
        self.file_path = file_path
        self.read_only = read_only
        self.data_only = data_only
        self.workbook: Optional[Workbook] = None
        self.cell_value_extractor = OpenpyxlCellValueExtractor()
    
    def open_workbook(self) -> None:
        """
        Open the Excel workbook for reading.
        
        Raises:
            FileNotFoundError: If the file does not exist
            ExcelAccessError: For other Excel access errors
        """
        if not os.path.exists(self.file_path):
            raise FileNotFoundError(f"File not found: {self.file_path}")
        
        try:
            self.workbook = openpyxl.load_workbook(
                self.file_path, 
                read_only=self.read_only, 
                data_only=self.data_only
            )
            logger.debug(f"Opened workbook: {self.file_path}")
        except Exception as e:
            logger.error(f"Error opening workbook: {str(e)}")
            raise ExcelAccessError(f"Error opening workbook: {str(e)}") from e
    
    def close_workbook(self) -> None:
        """
        Close the workbook and release resources.
        """
        if self.workbook:
            self.workbook.close()
            self.workbook = None
            logger.debug(f"Closed workbook: {self.file_path}")
    
    def get_sheet_names(self) -> List[str]:
        """
        Get all sheet names in the workbook.
        
        Returns:
            List of sheet names
            
        Raises:
            ExcelAccessError: If the workbook is not open
        """
        if not self.workbook:
            raise ExcelAccessError("Workbook is not open")
        
        return self.workbook.sheetnames
    
    def get_sheet_accessor(self, sheet_name: Optional[str] = None) -> 'OpenpyxlSheetAccessor':
        """
        Get a sheet accessor for the specified sheet.
        
        Args:
            sheet_name: Name of the sheet to access. If None, returns the active sheet.
            
        Returns:
            OpenpyxlSheetAccessor for the specified sheet
            
        Raises:
            ExcelAccessError: If the workbook is not open
            SheetNotFoundError: If the specified sheet does not exist
        """
        if not self.workbook:
            raise ExcelAccessError("Workbook is not open")
        
        try:
            if sheet_name is None:
                sheet = self.workbook.active
            else:
                if sheet_name not in self.workbook.sheetnames:
                    raise SheetNotFoundError(f"Sheet not found: {sheet_name}")
                
                sheet = self.workbook[sheet_name]
            
            return OpenpyxlSheetAccessor(sheet, self.cell_value_extractor)
        except Exception as e:
            if isinstance(e, SheetNotFoundError):
                raise
            
            logger.error(f"Error accessing sheet: {str(e)}")
            raise ExcelAccessError(f"Error accessing sheet: {str(e)}") from e


class OpenpyxlSheetAccessor(SheetAccessorInterface):
    """Sheet accessor implementation using openpyxl."""
    
    def __init__(self, sheet: Worksheet, cell_value_extractor: 'OpenpyxlCellValueExtractor'):
        """
        Initialize the sheet accessor.
        
        Args:
            sheet: Openpyxl worksheet
            cell_value_extractor: Extractor for typed cell values
        """
        self.sheet = sheet
        self.cell_value_extractor = cell_value_extractor
    
    @property
    def title(self) -> str:
        """Get the title of the underlying sheet."""
        return self.sheet.title
    
    def get_dimensions(self) -> Tuple[int, int, int, int]:
        """
        Get sheet dimensions as (min_row, max_row, min_col, max_col).
        
        Returns:
            Tuple of (min_row, max_row, min_col, max_col)
            
        Raises:
            ExcelAccessError: If dimensions cannot be determined
        """
        try:
            # Use the standard dimensions property from openpyxl
            min_row, min_col, max_row, max_col = openpyxl.utils.range_boundaries(self.sheet.dimensions)

            # Handle potentially empty sheets where dimensions might be None or invalid
            # openpyxl often returns (1, 1, 1, 1) for empty sheets after load, but we double-check
            if not all(isinstance(i, int) for i in [min_row, min_col, max_row, max_col]) or max_row < min_row or max_col < min_col:
                # Check if the sheet actually has any cells
                if not list(self.sheet.iter_rows(max_row=1, max_col=1)): # Check if A1 exists
                    return (1, 1, 1, 1) # Return default for truly empty sheet
                else:
                    # If sheet is not empty but dimensions are weird, raise error
                    raise ExcelAccessError("Could not determine valid sheet dimensions")

            return (min_row, max_row, min_col, max_col)
        except Exception as e:
            logger.error(f"Error getting sheet dimensions: {str(e)}")
            raise ExcelAccessError(f"Error getting sheet dimensions: {str(e)}") from e
    
    def get_merged_regions(self) -> List[Tuple[int, int, int, int]]:
        """
        Get all merged regions as (top_row, left_col, bottom_row, right_col) tuples.
        
        Returns:
            List of merged regions
            
        Raises:
            ExcelAccessError: If merged regions cannot be retrieved
        """
        try:
            merged_regions = []
            
            for merged_cell_range in self.sheet.merged_cells.ranges:
                min_row, min_col, max_row, max_col = merged_cell_range.bounds
                merged_regions.append((min_row, min_col, max_row, max_col))
            
            return merged_regions
        except Exception as e:
            logger.error(f"Error getting merged regions: {str(e)}")
            raise ExcelAccessError(f"Error getting merged regions: {str(e)}") from e
    
    def get_cell_value(self, row: int, column: int) -> Any:
        """
        Get the value of a cell with appropriate typing.
        
        Args:
            row: Row index (1-based)
            column: Column index (1-based)
            
        Returns:
            Typed cell value
            
        Raises:
            CellAccessError: If cell cannot be accessed
        """
        try:
            cell = self.sheet.cell(row=row, column=column)
            return cell.value
        except Exception as e:
            logger.error(f"Error accessing cell ({row}, {column}): {str(e)}")
            raise CellAccessError(f"Error accessing cell ({row}, {column}): {str(e)}") from e
    
    def get_row_values(self, row: int) -> Dict[int, Any]:
        """
        Get all values in a row as {column_index: value} dictionary.
        
        Args:
            row: Row index (1-based)
            
        Returns:
            Dictionary mapping column indices to cell values
            
        Raises:
            RowAccessError: If row cannot be accessed
        """
        try:
            row_values = {}
            
            # Get the dimensions to know the column range
            _, _, min_col, max_col = self.get_dimensions()
            
            for col in range(min_col, max_col + 1):
                cell = self.sheet.cell(row=row, column=col)
                if cell.value is not None:
                    row_values[col] = cell.value
            
            return row_values
        except Exception as e:
            if isinstance(e, ExcelAccessError):
                raise
            
            logger.error(f"Error accessing row {row}: {str(e)}")
            raise RowAccessError(f"Error accessing row {row}: {str(e)}") from e
    
    def iterate_rows(
        self, start_row: int, end_row: Optional[int] = None, 
        chunk_size: int = 1000
    ) -> Iterator[Dict[int, Dict[int, Any]]]:
        """
        Iterate through rows with chunking support.
        
        Args:
            start_row: First row to include (1-based)
            end_row: Last row to include (1-based), or None for all rows
            chunk_size: Number of rows to process in each iteration
            
        Yields:
            Dictionary mapping row indices to dictionaries of column values
            
        Raises:
            ExcelAccessError: If rows cannot be iterated
        """
        try:
            # Determine the end row if not specified
            if end_row is None:
                _, max_row, _, _ = self.get_dimensions()
                end_row = max_row
            
            # Process rows in chunks
            for chunk_start in range(start_row, end_row + 1, chunk_size):
                chunk_end = min(chunk_start + chunk_size - 1, end_row)
                
                logger.debug(f"Processing rows {chunk_start} to {chunk_end}")
                
                chunk_data = {}
                for row in range(chunk_start, chunk_end + 1):
                    row_values = self.get_row_values(row)
                    if row_values:  # Only include non-empty rows
                        chunk_data[row] = row_values
                
                yield chunk_data
        except Exception as e:
            if isinstance(e, (ExcelAccessError, RowAccessError)):
                raise
            
            logger.error(f"Error iterating rows: {str(e)}")
            raise ExcelAccessError(f"Error iterating rows: {str(e)}") from e


class OpenpyxlCellValueExtractor(CellValueExtractorInterface[Cell]):
    """Cell value extractor implementation for openpyxl."""
    
    def extract_string(self, value: Any) -> str:
        """
        Extract string value.
        
        Args:
            value: Cell value to convert
            
        Returns:
            String representation of the value
        """
        if value is None:
            return ""
        
        return str(value)
    
    def extract_number(self, value: Any) -> float:
        """
        Extract numeric value.
        
        Args:
            value: Cell value to convert
            
        Returns:
            Numeric representation of the value
            
        Raises:
            TypeError: If value cannot be converted to a number
        """
        if value is None:
            return 0.0
        
        try:
            return float(value)
        except (ValueError, TypeError):
            raise TypeError(f"Cannot convert {value} to a number")
    
    def extract_date(self, value: Any) -> str:
        """
        Extract date value as ISO format string.
        
        Args:
            value: Cell value to convert
            
        Returns:
            ISO-8601 formatted date string
            
        Raises:
            TypeError: If value cannot be converted to a date
        """
        if value is None:
            return ""
        
        if isinstance(value, (datetime, date)):
            return value.isoformat()
        
        raise TypeError(f"Cannot convert {value} to a date")
    
    def extract_boolean(self, value: Any) -> bool:
        """
        Extract boolean value.
        
        Args:
            value: Cell value to convert
            
        Returns:
            Boolean representation of the value
        """
        if value is None:
            return False
        
        if isinstance(value, bool):
            return value
        
        # Convert string representations
        if isinstance(value, str):
            value = value.lower()
            if value in ('true', 'yes', '1'):
                return True
            if value in ('false', 'no', '0'):
                return False
        
        # Convert numeric representations
        try:
            num_value = float(value)
            return bool(num_value)
        except (ValueError, TypeError):
            return bool(value)
    
    def detect_type(self, value: Any) -> str:
        """
        Detect the data type of a cell value.
        
        Args:
            value: Cell value to analyze
            
        Returns:
            String representing the detected type
        """
        if value is None:
            return 'null'
        
        if isinstance(value, bool):
            return 'boolean'
        
        if isinstance(value, (int, float)):
            return 'number'
        
        if isinstance(value, (datetime, date)):
            return 'date'
        
        return 'string'


class ExcelAccessError(Exception):
    """Base exception for Excel access errors."""
    pass


class SheetNotFoundError(ExcelAccessError):
    """Exception raised when a sheet is not found."""
    pass


class CellAccessError(ExcelAccessError):
    """Exception raised when a cell cannot be accessed."""
    pass


class RowAccessError(ExcelAccessError):
    """Exception raised when a row cannot be accessed."""
    pass


class UnsupportedFileError(ExcelAccessError):
    """Exception raised when a file cannot be handled by the strategy."""
    pass
