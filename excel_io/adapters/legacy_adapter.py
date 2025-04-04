import logging
from typing import Any, Dict, Iterator, List, Optional, Tuple

from ..interfaces import ExcelReaderInterface, SheetAccessorInterface, CellValueExtractorInterface

logger = logging.getLogger(__name__)


class LegacyReaderAdapter(ExcelReaderInterface):
    """
    Adapter for legacy Excel reader implementations.
    
    This adapter allows existing code that uses the legacy reader
    to use the new ExcelReaderInterface.
    """
    
    def __init__(self, legacy_reader):
        """
        Initialize the adapter with a legacy reader.
        
        Args:
            legacy_reader: Legacy Excel reader instance
        """
        self.legacy_reader = legacy_reader
        self.cell_value_extractor = LegacyCellValueExtractor()
    
    def open_workbook(self) -> None:
        """
        Open the Excel workbook using the legacy reader.
        
        Raises:
            ExcelAccessError: If the workbook cannot be opened
        """
        try:
            # Legacy readers often don't have explicit open methods,
            # or the file is opened in the constructor
            if hasattr(self.legacy_reader, 'open'):
                self.legacy_reader.open()
            
            logger.debug("Opened workbook using legacy reader")
        except Exception as e:
            logger.error(f"Error opening workbook with legacy reader: {str(e)}")
            raise ExcelAccessError(f"Error opening workbook with legacy reader: {str(e)}") from e
    
    def close_workbook(self) -> None:
        """
        Close the workbook using the legacy reader.
        """
        try:
            if hasattr(self.legacy_reader, 'close'):
                self.legacy_reader.close()
            
            logger.debug("Closed workbook using legacy reader")
        except Exception as e:
            logger.warning(f"Error closing workbook with legacy reader: {str(e)}")
    
    def get_sheet_names(self) -> List[str]:
        """
        Get all sheet names using the legacy reader.
        
        Returns:
            List of sheet names
            
        Raises:
            ExcelAccessError: If sheet names cannot be retrieved
        """
        try:
            if hasattr(self.legacy_reader, 'get_sheet_names'):
                return self.legacy_reader.get_sheet_names()
            
            # Alternative method for legacy readers
            if hasattr(self.legacy_reader, 'wb') and hasattr(self.legacy_reader.wb, 'sheetnames'):
                return self.legacy_reader.wb.sheetnames
            
            # Last resort
            if hasattr(self.legacy_reader, 'workbook') and hasattr(self.legacy_reader.workbook, 'sheetnames'):
                return self.legacy_reader.workbook.sheetnames
            
            raise ExcelAccessError("Cannot get sheet names from legacy reader")
        except Exception as e:
            logger.error(f"Error getting sheet names from legacy reader: {str(e)}")
            raise ExcelAccessError(f"Error getting sheet names from legacy reader: {str(e)}") from e
    
    def get_sheet_accessor(self, sheet_name: Optional[str] = None) -> 'LegacySheetAdapter':
        """
        Get a sheet accessor using the legacy reader.
        
        Args:
            sheet_name: Name of the sheet to access. If None, returns the active sheet.
            
        Returns:
            LegacySheetAdapter for the specified sheet
            
        Raises:
            SheetNotFoundError: If the specified sheet does not exist
            ExcelAccessError: For other Excel access errors
        """
        try:
            if hasattr(self.legacy_reader, 'get_sheet'):
                legacy_sheet = self.legacy_reader.get_sheet(sheet_name)
            elif hasattr(self.legacy_reader, 'wb'):
                if sheet_name is None:
                    legacy_sheet = self.legacy_reader.wb.active
                else:
                    try:
                        legacy_sheet = self.legacy_reader.wb[sheet_name]
                    except KeyError:
                        raise SheetNotFoundError(f"Sheet not found: {sheet_name}")
            elif hasattr(self.legacy_reader, 'workbook'):
                if sheet_name is None:
                    legacy_sheet = self.legacy_reader.workbook.active
                else:
                    try:
                        legacy_sheet = self.legacy_reader.workbook[sheet_name]
                    except KeyError:
                        raise SheetNotFoundError(f"Sheet not found: {sheet_name}")
            else:
                raise ExcelAccessError("Cannot get sheet from legacy reader")
            
            return LegacySheetAdapter(legacy_sheet, self.cell_value_extractor)
        except Exception as e:
            if isinstance(e, SheetNotFoundError):
                raise
            
            logger.error(f"Error getting sheet from legacy reader: {str(e)}")
            raise ExcelAccessError(f"Error getting sheet from legacy reader: {str(e)}") from e


class LegacySheetAdapter(SheetAccessorInterface):
    """
    Adapter for legacy Excel sheet implementations.
    
    This adapter allows existing code that uses legacy sheet objects
    to use the new SheetAccessorInterface.
    """
    
    def __init__(self, legacy_sheet, cell_value_extractor: 'LegacyCellValueExtractor'):
        """
        Initialize the adapter with a legacy sheet.
        
        Args:
            legacy_sheet: Legacy Excel sheet object
            cell_value_extractor: Extractor for typed cell values
        """
        self.legacy_sheet = legacy_sheet
        self.cell_value_extractor = cell_value_extractor
    
    def get_dimensions(self) -> Tuple[int, int, int, int]:
        """
        Get sheet dimensions from the legacy sheet.
        
        Returns:
            Tuple of (min_row, max_row, min_col, max_col)
            
        Raises:
            ExcelAccessError: If dimensions cannot be determined
        """
        try:
            if hasattr(self.legacy_sheet, 'dimensions'):
                # Try to parse dimensions string like "A1:Z100"
                if isinstance(self.legacy_sheet.dimensions, str):
                    parts = self.legacy_sheet.dimensions.split(':')
                    if len(parts) == 2:
                        from openpyxl.utils import coordinate_from_string, column_index_from_string
                        start_coord = coordinate_from_string(parts[0])
                        end_coord = coordinate_from_string(parts[1])
                        
                        min_col = column_index_from_string(start_coord[0])
                        min_row = start_coord[1]
                        max_col = column_index_from_string(end_coord[0])
                        max_row = end_coord[1]
                        
                        return (min_row, max_row, min_col, max_col)
            
            # Try direct properties
            if (hasattr(self.legacy_sheet, 'min_row') and 
                hasattr(self.legacy_sheet, 'max_row') and 
                hasattr(self.legacy_sheet, 'min_column') and 
                hasattr(self.legacy_sheet, 'max_column')):
                return (
                    self.legacy_sheet.min_row or 1, 
                    self.legacy_sheet.max_row or 1,
                    self.legacy_sheet.min_column or 1, 
                    self.legacy_sheet.max_column or 1
                )
            
            # Extract from data range (common pattern)
            return (1, 1, 1, 1)  # Default fallback
        except Exception as e:
            logger.error(f"Error getting dimensions from legacy sheet: {str(e)}")
            raise ExcelAccessError(f"Error getting dimensions from legacy sheet: {str(e)}") from e
    
    def get_merged_regions(self) -> List[Tuple[int, int, int, int]]:
        """
        Get merged regions from the legacy sheet.
        
        Returns:
            List of merged regions as (top, left, bottom, right) tuples
            
        Raises:
            ExcelAccessError: If merged regions cannot be retrieved
        """
        try:
            merged_regions = []
            
            # Try different property names from common legacy readers
            if hasattr(self.legacy_sheet, 'merged_cells') and hasattr(self.legacy_sheet.merged_cells, 'ranges'):
                # openpyxl style
                for merged_cell_range in self.legacy_sheet.merged_cells.ranges:
                    if hasattr(merged_cell_range, 'bounds'):
                        merged_regions.append(merged_cell_range.bounds)
                    elif hasattr(merged_cell_range, 'min_row'):
                        merged_regions.append((
                            merged_cell_range.min_row,
                            merged_cell_range.min_col,
                            merged_cell_range.max_row,
                            merged_cell_range.max_col
                        ))
            elif hasattr(self.legacy_sheet, 'merged_ranges'):
                # Another common pattern
                for merged_range in self.legacy_sheet.merged_ranges:
                    if isinstance(merged_range, str):
                        # Parse range string like "A1:B2"
                        from openpyxl.utils import range_boundaries
                        bounds = range_boundaries(merged_range)
                        merged_regions.append(bounds)
            
            return merged_regions
        except Exception as e:
            logger.error(f"Error getting merged regions from legacy sheet: {str(e)}")
            raise ExcelAccessError(f"Error getting merged regions from legacy sheet: {str(e)}") from e
    
    def get_cell_value(self, row: int, column: int) -> Any:
        """
        Get cell value from the legacy sheet.
        
        Args:
            row: Row index (1-based)
            column: Column index (1-based)
            
        Returns:
            Cell value
            
        Raises:
            CellAccessError: If cell cannot be accessed
        """
        try:
            # Try different ways to access cells
            if hasattr(self.legacy_sheet, 'cell'):
                cell = self.legacy_sheet.cell(row=row, column=column)
                return cell.value
            
            # Try dictionary-style access
            try:
                from openpyxl.utils import get_column_letter
                cell_coord = f"{get_column_letter(column)}{row}"
                return self.legacy_sheet[cell_coord].value
            except (ImportError, AttributeError, KeyError):
                pass
            
            # Last resort: try direct 2D indexing
            try:
                return self.legacy_sheet[row-1][column-1]
            except (IndexError, TypeError):
                pass
            
            # Give up
            return None
        except Exception as e:
            logger.error(f"Error getting cell value from legacy sheet: {str(e)}")
            raise CellAccessError(f"Error getting cell value from legacy sheet: {str(e)}") from e
    
    def get_row_values(self, row: int) -> Dict[int, Any]:
        """
        Get row values from the legacy sheet.
        
        Args:
            row: Row index (1-based)
            
        Returns:
            Dictionary mapping column indices to cell values
            
        Raises:
            RowAccessError: If row cannot be accessed
        """
        try:
            row_values = {}
            
            # Try different ways to access rows
            if hasattr(self.legacy_sheet, 'iter_rows'):
                for cell in next(self.legacy_sheet.iter_rows(min_row=row, max_row=row), []):
                    if cell.value is not None:
                        row_values[cell.column] = cell.value
            else:
                # Extract row using dimensions
                _, _, min_col, max_col = self.get_dimensions()
                for col in range(min_col, max_col + 1):
                    value = self.get_cell_value(row, col)
                    if value is not None:
                        row_values[col] = value
            
            return row_values
        except Exception as e:
            logger.error(f"Error getting row values from legacy sheet: {str(e)}")
            raise RowAccessError(f"Error getting row values from legacy sheet: {str(e)}") from e
    
    def iterate_rows(
        self, start_row: int, end_row: Optional[int] = None, 
        chunk_size: int = 1000
    ) -> Iterator[Dict[int, Dict[int, Any]]]:
        """
        Iterate through rows from the legacy sheet.
        
        Args:
            start_row: First row to include (1-based)
            end_row: Last row to include (1-based), or None for all rows
            chunk_size: Number of rows to process in each iteration
            
        Yields:
            Dictionary mapping row indices to dictionaries of column values
            
        Raises:
            ExcelAccessError: If rows cannot be iterated
        """
        try:
            # Determine the end row if not specified
            if end_row is None:
                _, max_row, _, _ = self.get_dimensions()
                end_row = max_row
            
            # Process rows in chunks
            for chunk_start in range(start_row, end_row + 1, chunk_size):
                chunk_end = min(chunk_start + chunk_size - 1, end_row)
                
                chunk_data = {}
                for row in range(chunk_start, chunk_end + 1):
                    row_values = self.get_row_values(row)
                    if row_values:  # Only include non-empty rows
                        chunk_data[row] = row_values
                
                yield chunk_data
        except Exception as e:
            logger.error(f"Error iterating rows from legacy sheet: {str(e)}")
            raise ExcelAccessError(f"Error iterating rows from legacy sheet: {str(e)}") from e


class LegacyCellValueExtractor(CellValueExtractorInterface[Any]):
    """
    Cell value extractor for legacy cell values.
    """
    
    def extract_string(self, value: Any) -> str:
        """
        Extract string value.
        
        Args:
            value: Cell value to convert
            
        Returns:
            String representation of the value
        """
        if value is None:
            return ""
        
        return str(value)
    
    def extract_number(self, value: Any) -> float:
        """
        Extract numeric value.
        
        Args:
            value: Cell value to convert
            
        Returns:
            Numeric representation of the value
            
        Raises:
            TypeError: If value cannot be converted to a number
        """
        if value is None:
            return 0.0
        
        try:
            return float(value)
        except (ValueError, TypeError):
            raise TypeError(f"Cannot convert {value} to a number")
    
    def extract_date(self, value: Any) -> str:
        """
        Extract date value as ISO format string.
        
        Args:
            value: Cell value to convert
            
        Returns:
            ISO-8601 formatted date string
            
        Raises:
            TypeError: If value cannot be converted to a date
        """
        if value is None:
            return ""
        
        from datetime import datetime, date
        
        if isinstance(value, (datetime, date)):
            return value.isoformat()
        
        raise TypeError(f"Cannot convert {value} to a date")
    
    def extract_boolean(self, value: Any) -> bool:
        """
        Extract boolean value.
        
        Args:
            value: Cell value to convert
            
        Returns:
            Boolean representation of the value
        """
        if value is None:
            return False
        
        if isinstance(value, bool):
            return value
        
        # Handle string values
        if isinstance(value, str):
            value = value.lower()
            if value in ('true', 'yes', '1', 't', 'y'):
                return True
            if value in ('false', 'no', '0', 'f', 'n'):
                return False
        
        # Handle numeric values
        try:
            num_value = float(value)
            return bool(num_value)
        except (ValueError, TypeError):
            return bool(value)
    
    def detect_type(self, value: Any) -> str:
        """
        Detect the data type of a cell value.
        
        Args:
            value: Cell value to analyze
            
        Returns:
            String representing the detected type
        """
        if value is None:
            return 'null'
        
        if isinstance(value, bool):
            return 'boolean'
        
        if isinstance(value, (int, float)):
            return 'number'
        
        from datetime import datetime, date
        if isinstance(value, (datetime, date)):
            return 'date'
        
        return 'string'


class ExcelAccessError(Exception):
    """Base exception for Excel access errors."""
    pass


class SheetNotFoundError(ExcelAccessError):
    """Exception raised when a sheet is not found."""
    pass


class CellAccessError(ExcelAccessError):
    """Exception raised when a cell cannot be accessed."""
    pass


class RowAccessError(ExcelAccessError):
    """Exception raised when a row cannot be accessed."""
    pass
