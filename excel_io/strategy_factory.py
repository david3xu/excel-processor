import os
import logging
import tempfile
from typing import Dict, List, Optional, Set, Type, Any

from openpyxl import load_workbook

from .interfaces import ExcelReaderInterface
from .strategies.base_strategy import ExcelAccessStrategy

logger = logging.getLogger(__name__)


class StrategyFactory:
    """Factory for creating Excel access strategies based on file characteristics."""
    
    def __init__(self, config: Optional[Dict[str, Any]] = None):
        """
        Initialize the strategy factory.
        
        Args:
            config: Configuration dictionary for strategy selection
        """
        self.strategies: List[ExcelAccessStrategy] = []
        self.config = config or {}
        self._preferred_strategy = self.config.get("preferred_strategy", "auto")
        self._enable_fallback = self.config.get("enable_fallback", True)
        self._large_file_threshold_mb = self.config.get("large_file_threshold_mb", 50)
        self._complex_structure_detection = self.config.get("complex_structure_detection", True)
    
    def register_strategy(self, strategy: ExcelAccessStrategy) -> None:
        """
        Register a strategy with the factory.
        
        Args:
            strategy: ExcelAccessStrategy implementation to register
        """
        self.strategies.append(strategy)
        logger.debug(f"Registered strategy: {strategy.get_strategy_name()}")
    
    def create_reader(self, file_path: str, **kwargs) -> ExcelReaderInterface:
        """
        Create a reader for the specified Excel file using the optimal strategy.
        
        Args:
            file_path: Path to the Excel file
            **kwargs: Additional parameters to pass to the strategy
            
        Returns:
            ExcelReaderInterface implementation for the specified file
            
        Raises:
            FileNotFoundError: If the file does not exist
            UnsupportedFileError: If no strategy can handle the file
            ExcelAccessError: For other Excel access errors
        """
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"File not found: {file_path}")
        
        # Determine the optimal strategy for this file
        strategy = self.determine_optimal_strategy(file_path)
        if strategy is None:
            raise UnsupportedFileError(f"No strategy can handle this file: {file_path}")
        
        logger.info(f"Using strategy {strategy.get_strategy_name()} for file: {file_path}")
        try:
            return strategy.create_reader(file_path, **kwargs)
        except Exception as e:
            if self._enable_fallback:
                return self._try_fallback_strategies(file_path, strategy, e, **kwargs)
            raise
    
    def determine_optimal_strategy(self, file_path: str) -> Optional[ExcelAccessStrategy]:
        """
        Determine the optimal strategy for the specified file.
        
        Args:
            file_path: Path to the Excel file
            
        Returns:
            The optimal ExcelAccessStrategy implementation, or None if no strategy can handle the file
        """
        # If a preferred strategy is specified and not "auto", try it first
        if self._preferred_strategy != "auto":
            for strategy in self.strategies:
                if strategy.get_strategy_name().lower() == self._preferred_strategy.lower():
                    if strategy.can_handle_file(file_path):
                        return strategy
        
        # Check file characteristics
        file_size_mb = os.path.getsize(file_path) / (1024 * 1024)
        is_large_file = file_size_mb > self._large_file_threshold_mb
        
        # Check for complex structures if enabled
        has_complex_structure = False
        if self._complex_structure_detection:
            has_complex_structure = self._detect_complex_structure(file_path)
        
        # For large files, prefer pandas strategy if available
        if is_large_file and not has_complex_structure:
            for strategy in self.strategies:
                if (strategy.get_strategy_name().lower() == "pandas" and 
                    strategy.can_handle_file(file_path)):
                    return strategy
        
        # For complex structures, prefer openpyxl strategy if available
        if has_complex_structure:
            for strategy in self.strategies:
                if (strategy.get_strategy_name().lower() == "openpyxl" and 
                    strategy.can_handle_file(file_path)):
                    return strategy
        
        # Try all other strategies in order
        for strategy in self.strategies:
            if strategy.can_handle_file(file_path):
                return strategy
        
        return None
    
    def _detect_complex_structure(self, file_path: str) -> bool:
        """
        Detect if the Excel file has complex structures requiring specialized handling.
        
        Complex structures include:
        - Merged cells
        - Formulas referencing multiple sheets
        - Conditional formatting
        - Data validation rules
        - Pivot tables
        
        Args:
            file_path: Path to the Excel file
            
        Returns:
            True if complex structures detected, False otherwise
        """
        try:
            # Use openpyxl directly here for structure analysis
            # We need to use read_only=False to access merged cells
            with tempfile.TemporaryDirectory() as tmp_dir:
                # Use a temporary directory to avoid potential file locking issues
                wb = load_workbook(file_path, read_only=False, data_only=False, keep_vba=False)
                
                # Check for merged cells in any sheet
                for sheet_name in wb.sheetnames:
                    sheet = wb[sheet_name]
                    if sheet.merged_cells and len(sheet.merged_cells.ranges) > 0:
                        logger.debug(f"Detected merged cells in sheet {sheet_name}")
                        return True
                
                # Check for complex formulas
                formula_count = 0
                reference_sheets = set()
                
                for sheet_name in wb.sheetnames:
                    sheet = wb[sheet_name]
                    for row in sheet.iter_rows(min_row=1, max_row=min(100, sheet.max_row or 1)):
                        for cell in row:
                            if cell.data_type == 'f':  # Formula
                                formula_count += 1
                                # Check for cross-sheet references
                                formula = str(cell.value).upper()
                                if "!" in formula:
                                    reference_sheets.add(formula.split("!")[0].strip("'"))
                
                # If many formulas or cross-sheet references
                if formula_count > 10 or len(reference_sheets) > 1:
                    logger.debug(f"Detected complex formulas: {formula_count} formulas, {len(reference_sheets)} referenced sheets")
                    return True
                
                # Check for conditional formatting
                for sheet_name in wb.sheetnames:
                    sheet = wb[sheet_name]
                    if hasattr(sheet, 'conditional_formatting') and sheet.conditional_formatting:
                        logger.debug(f"Detected conditional formatting in sheet {sheet_name}")
                        return True
                
                # If we get here, no complex structures were found
                return False
                
        except Exception as e:
            logger.warning(f"Error during complex structure detection: {str(e)}")
            # On error, assume no complex structures
            return False
    
    def _try_fallback_strategies(
        self, file_path: str, failed_strategy: ExcelAccessStrategy, 
        original_error: Exception, **kwargs
    ) -> ExcelReaderInterface:
        """
        Try fallback strategies after a failure.
        
        Args:
            file_path: Path to the Excel file
            failed_strategy: The strategy that failed
            original_error: The exception raised by the failed strategy
            **kwargs: Additional parameters to pass to the strategy
            
        Returns:
            ExcelReaderInterface implementation from a fallback strategy
            
        Raises:
            The original exception if no fallback strategy succeeds
        """
        logger.warning(
            f"Strategy {failed_strategy.get_strategy_name()} failed with error: {str(original_error)}. "
            f"Trying fallback strategies."
        )
        
        for strategy in self.strategies:
            if strategy is not failed_strategy and strategy.can_handle_file(file_path):
                try:
                    logger.info(f"Trying fallback strategy: {strategy.get_strategy_name()}")
                    return strategy.create_reader(file_path, **kwargs)
                except Exception as e:
                    logger.warning(f"Fallback strategy {strategy.get_strategy_name()} "
                                 f"also failed: {str(e)}")
        
        logger.error("All fallback strategies failed")
        raise original_error


class UnsupportedFileError(Exception):
    """Exception raised when no strategy can handle a file."""
    pass

