"""
Data extractor for Excel files.
Extracts hierarchical data while respecting merged cells and structure.
"""

from typing import Any, Dict, List, Optional, Set, Tuple, Union, Generator, Callable

import os
import psutil
import openpyxl
import pandas as pd
from openpyxl.worksheet.worksheet import Worksheet
from pydantic import BaseModel, Field

from core.reader import ExcelReader, RowData, WorksheetData
from models.excel_structure import (CellPosition, SheetStructure)
from models.hierarchical_data import (HierarchicalData,
                                                    HierarchicalDataItem,
                                                    HierarchicalRecord,
                                                    MergeInfo)
from utils.exceptions import DataExtractionError, HierarchicalDataError, MemoryError
from utils.logging import get_logger
from utils.performance import selective_validation, StreamingValidator

logger = get_logger(__name__)


class ExtractionError(Exception):
    """Exception raised during data extraction."""
    pass


class DataChunk(BaseModel):
    """
    Model for a chunk of data extracted during streaming processing.
    
    Attributes:
        chunk_index: Index of this chunk in the overall extraction
        rows: List of row data in this chunk
        is_final: Whether this is the final chunk of data
    """
    chunk_index: int = Field(..., description="Index of this chunk in the overall extraction")
    rows: List[RowData] = Field(default_factory=list, description="List of row data in this chunk")
    is_final: bool = Field(False, description="Whether this is the final chunk of data")


class StreamingDataExtractor:
    """
    Extracts data from Excel files in a memory-efficient streaming manner.
    
    This class provides optimized extraction of data from large Excel files
    with memory usage optimization and validation controls for performance.
    
    Attributes:
        chunk_size: Number of rows to process in each chunk
        validator: Streaming validator for performance optimization
    """
    
    def __init__(self, chunk_size: int = 1000):
        """
        Initialize the streaming data extractor.
        
        Args:
            chunk_size: Number of rows to process in each chunk
        """
        self.chunk_size = chunk_size
        self.validator = StreamingValidator(
            model_class=RowData,
            validation_interval=20  # Validate every 20th row
        )
    
    def extract_from_worksheet(
        self,
        reader: Any,
        sheet: Any,
        min_row: int = 1,
        max_row: Optional[int] = None,
        skip_empty: bool = True
    ) -> Generator[DataChunk, None, None]:
        """
        Extract data from a worksheet in chunks.
        
        This method streams data from a worksheet, processing it in chunks
        to optimize memory usage while applying validation controls.
        
        Args:
            reader: Excel reader instance
            sheet: Worksheet object from the reader
            min_row: Minimum row index to include (1-based)
            max_row: Maximum row index to include (1-based)
            skip_empty: Whether to skip empty rows
            
        Yields:
            DataChunk objects containing processed rows
        """
        chunk_index = 0
        has_more_chunks = True
        
        while has_more_chunks:
            # Process rows in chunks using the reader's iter_rows method
            try:
                rows_processed = 0
                chunk_rows = []
                
                # Get the next chunk of rows
                for row_batch in reader.iter_rows(
                    sheet,
                    min_row=min_row + (chunk_index * self.chunk_size),
                    max_row=min(
                        min_row + ((chunk_index + 1) * self.chunk_size - 1),
                        max_row or float('inf')
                    ),
                    skip_empty=skip_empty,
                    chunk_size=self.chunk_size
                ):
                    # Add rows to the current chunk
                    chunk_rows.extend(row_batch)
                    rows_processed += len(row_batch)
                    
                    # Stop if we've reached our chunk size
                    if rows_processed >= self.chunk_size:
                        break
                
                # Check if we've processed any rows
                if not chunk_rows:
                    # No more rows to process
                    has_more_chunks = False
                    continue
                
                # Create and yield the data chunk
                is_final = False
                if (
                    (max_row is not None and min_row + ((chunk_index + 1) * self.chunk_size) > max_row) or
                    rows_processed < self.chunk_size
                ):
                    is_final = True
                    has_more_chunks = False
                
                yield DataChunk(
                    chunk_index=chunk_index,
                    rows=chunk_rows,
                    is_final=is_final
                )
                
                # Move to the next chunk
                chunk_index += 1
                min_row += rows_processed
            
            except Exception as e:
                logger.error(f"Error extracting data: {e}")
                raise ExtractionError(f"Failed to extract data from worksheet: {str(e)}") from e
    
    def process_worksheet(
        self,
        reader: Any,
        sheet: Any,
        output_handler: Optional[Callable[[DataChunk], None]] = None,
        skip_empty: bool = True,
        min_row: int = 1,
        max_row: Optional[int] = None
    ) -> Dict[str, Any]:
        """
        Process an entire worksheet with streaming extraction.
        
        This method processes a worksheet by extracting data in chunks and
        optionally passing each chunk to an output handler function.
        
        Args:
            reader: Excel reader instance
            sheet: Worksheet object from the reader
            output_handler: Function to call with each data chunk
            skip_empty: Whether to skip empty rows
            min_row: Minimum row index to include (1-based)
            max_row: Maximum row index to include (1-based)
            
        Returns:
            Dictionary with processing statistics
        """
        chunk_count = 0
        row_count = 0
        
        # Begin extraction
        logger.info(f"Starting streaming extraction with chunk size {self.chunk_size}")
        
        # Get sheet name
        sheet_name = None
        if hasattr(sheet, 'title'):
            sheet_name = sheet.title
        elif isinstance(sheet, str):
            sheet_name = sheet
        
        for chunk in self.extract_from_worksheet(
            reader, sheet, min_row=min_row, max_row=max_row, skip_empty=skip_empty
        ):
            # Count rows and chunks
            chunk_count += 1
            row_count += len(chunk.rows)
            
            # Call output handler if provided
            if output_handler:
                output_handler(chunk)
            
            logger.info(f"Processed chunk {chunk_count} with {len(chunk.rows)} rows")
        
        # Return statistics
        return {
            "sheet_name": sheet_name,
            "chunk_count": chunk_count,
            "row_count": row_count,
            "status": "success"
        }


class DataExtractor:
    """
    Extracts data from workbooks using memory-efficient techniques.
    
    This class provides methods for extracting data from Excel workbooks
    and applying validation rules during the extraction process.
    """
    
    def __init__(self, performance_mode: bool = True):
        """
        Initialize the data extractor.
        
        Args:
            performance_mode: Whether to optimize for performance
        """
        self.performance_mode = performance_mode
    
    def extract_worksheet_data(
        self,
        reader: Any,
        sheet_name: Optional[str] = None,
        include_empty_rows: bool = False
    ) -> WorksheetData:
        """
        Extract data from a worksheet.
        
        This method extracts data from a worksheet and validates it
        according to our data models.
        
        Args:
            reader: Excel reader instance
            sheet_name: Name of the sheet to extract, or None for the first sheet
            include_empty_rows: Whether to include empty rows
            
        Returns:
            WorksheetData object with extracted and validated data
        """
        try:
            # Get the sheet
            sheet = reader.get_sheet(sheet_name)
            
            # Create worksheet model
            worksheet_data = reader.create_worksheet_model(
                sheet,
                include_empty_rows=include_empty_rows,
                performance_mode=self.performance_mode
            )
            
            logger.info(
                f"Extracted {worksheet_data.row_count} rows from sheet "
                f"{worksheet_data.name}"
            )
            
            return worksheet_data
        
        except Exception as e:
            logger.error(f"Error extracting worksheet data: {e}")
            raise ExtractionError(f"Failed to extract worksheet data: {str(e)}") from e
    
    def create_streaming_extractor(self, chunk_size: int = 1000) -> StreamingDataExtractor:
        """
        Create a streaming data extractor.
        
        This factory method creates a configured StreamingDataExtractor
        for processing large files efficiently.
        
        Args:
            chunk_size: Number of rows to process in each chunk
            
        Returns:
            Configured StreamingDataExtractor instance
        """
        return StreamingDataExtractor(chunk_size=chunk_size)

    def extract_data(
        self,
        sheet: Worksheet,
        merge_map: Dict[Tuple[int, int], Dict],
        data_start_row: int,
        chunk_size: int = 1000,
        include_empty: bool = False
    ) -> HierarchicalData:
        """
        Extract hierarchical data from an Excel sheet.
        
        Args:
            sheet: Worksheet to extract data from
            merge_map: Dictionary mapping (row, col) to merge information
            data_start_row: Row number where data starts (including header)
            chunk_size: Number of rows to process at once
            include_empty: Whether to include empty cells
            
        Returns:
            HierarchicalData instance with extracted data
            
        Raises:
            DataExtractionError: If data extraction fails
        """
        try:
            logger.info(
                f"Extracting hierarchical data from sheet: {sheet.title} "
                f"starting at row {data_start_row}"
            )
            
            # Get dimensions needed for header extraction
            _, _, min_col, max_col = sheet.get_dimensions()
            
            # Get header row data - fetch DIRECTLY to avoid merge issues for headers
            # header_row_data = sheet.get_row_values(data_start_row) # OLD way - affected by merges below
            header_map = {} # Recreate header_map (col_idx -> col_name) directly
            for col_idx in range(min_col, max_col + 1):
                # Use sheet accessor's underlying cell access if possible, or assume direct access for this specific case
                # This relies on the accessor having a way to get the raw cell or value at the specific row/col
                # Assuming get_cell_value fetches the value at the specific coordinate
                header_value = sheet.get_cell_value(data_start_row, col_idx)
                if header_value is not None:
                     header_map[col_idx] = str(header_value)
                     
            logger.debug(f"Directly fetched header map (data_start_row={data_start_row}): {header_map}")
            
            # Directly fetch header names from the cells in the header row
            headers = self._get_header_values_directly(sheet, data_start_row, header_map)
            logger.debug(f"Created headers list from direct fetch: {headers}")
            logger.debug(f"Extractor: headers list id: {id(headers)}")

            data = HierarchicalData(columns=headers)
            logger.debug(f"Extractor: data.columns id: {id(data.columns)}")
            
            # Determine total rows to process
            _, max_row, _, _ = sheet.get_dimensions()
            data_end_row = max_row
            total_rows_to_process = max(0, data_end_row - data_start_row)
            
            logger.info(f"Processing {total_rows_to_process} data rows (from {data_start_row + 1} to {data_end_row})")
            
            processed_rows_count = 0
            # Iterate through data rows using the accessor
            # Access the underlying openpyxl sheet object for iter_rows
            for row_chunk in self._iterate_rows_chunked(sheet, data_start_row + 1, data_end_row, chunk_size):
                # Process each row in the chunk
                for excel_row_idx, row_data in row_chunk.items():
                    # Process row and add to hierarchical data
                    record = self._process_row(
                        sheet, row_data, excel_row_idx, header_map, merge_map, include_empty
                    )
                    data.add_record(record)
                    processed_rows_count += 1
            
            logger.info(f"Extracted {len(data.records)} records from {processed_rows_count} processed rows.")
            return data
        except Exception as e:
            error_msg = f"Failed to extract hierarchical data: {str(e)}"
            logger.error(error_msg)
            raise DataExtractionError(error_msg) from e
    
    def extract_data_streaming(
        self,
        sheet: Worksheet,
        merge_map: Dict[Tuple[int, int], Dict],
        data_start_row: int,
        chunk_size: int = 1000,
        include_empty: bool = False,
        memory_threshold: float = 0.8
    ) -> Generator[Tuple[HierarchicalData, bool], None, None]:
        """
        Extract hierarchical data from an Excel sheet in a streaming fashion.
        Yields chunks of data rather than returning a complete dataset.
        
        Args:
            sheet: Worksheet to extract data from
            merge_map: Dictionary mapping (row, col) to merge information
            data_start_row: Row number where data starts (including header)
            chunk_size: Number of rows to process at once
            include_empty: Whether to include empty cells
            memory_threshold: Memory usage threshold (0.0-1.0) for dynamic chunk size adjustment
            
        Yields:
            Tuple containing:
              - HierarchicalData with the current chunk's data
              - Boolean indicating if this is the final chunk
              
        Raises:
            DataExtractionError: If data extraction fails
            MemoryError: If memory usage exceeds safe limits
        """
        try:
            logger.info(
                f"Streaming extraction of hierarchical data from sheet: {sheet.title} "
                f"starting at row {data_start_row} with initial chunk size {chunk_size}"
            )
            
            # Get dimensions needed for header extraction
            _, _, min_col, max_col = sheet.get_dimensions()
            
            # Get header map directly (same as non-streaming version)
            header_map = {}
            for col_idx in range(min_col, max_col + 1):
                header_value = sheet.get_cell_value(data_start_row, col_idx)
                if header_value is not None:
                     header_map[col_idx] = str(header_value)
            
            # Directly fetch header names from the cells in the header row
            headers = self._get_header_values_directly(sheet, data_start_row, header_map)
            logger.debug(f"Created headers list from direct fetch for streaming: {headers}")
            
            # Determine total rows to process
            _, max_row, _, _ = sheet.get_dimensions()
            data_end_row = max_row
            total_rows_to_process = max(0, data_end_row - data_start_row)
            
            logger.info(f"Streaming process for {total_rows_to_process} data rows (from {data_start_row + 1} to {data_end_row})")
            
            # Monitor memory usage to adjust chunk size
            process = psutil.Process(os.getpid())
            adaptive_chunk_size = chunk_size
            
            processed_rows_count = 0
            total_rows_yielded = 0
            
            # Iterate through data rows in chunks with adaptive chunk size
            for row_chunk in self._iterate_rows_chunked(sheet, data_start_row + 1, data_end_row, adaptive_chunk_size):
                # Create a new HierarchicalData instance for this chunk
                chunk_data = HierarchicalData(columns=headers)
                rows_in_current_chunk = 0
                
                # Process each row in the chunk
                for excel_row_idx, row_data in row_chunk.items():
                    # Process row and add to the chunk data
                    try:
                        record = self._process_row(
                            sheet, row_data, excel_row_idx, header_map, merge_map, include_empty
                        )
                        chunk_data.add_record(record)
                        processed_rows_count += 1
                        rows_in_current_chunk += 1
                    except Exception as e:
                        logger.error(f"Error processing row {excel_row_idx}: {str(e)}")
                        # Continue processing other rows despite errors in individual rows
                        continue
                
                total_rows_yielded += rows_in_current_chunk
                is_final_chunk = (total_rows_yielded >= total_rows_to_process)
                
                # Check memory usage and adjust chunk size for next iteration
                mem_percent = process.memory_percent()
                if mem_percent > memory_threshold:
                    # Reduce chunk size to avoid memory issues
                    adaptive_chunk_size = max(100, int(adaptive_chunk_size * 0.7))
                    logger.warning(
                        f"Memory usage high ({mem_percent:.1f}%), reducing chunk size to {adaptive_chunk_size}"
                    )
                elif mem_percent < (memory_threshold * 0.5) and adaptive_chunk_size < chunk_size:
                    # Increase chunk size if memory usage is low
                    adaptive_chunk_size = min(chunk_size, int(adaptive_chunk_size * 1.3))
                    logger.info(
                        f"Memory usage low ({mem_percent:.1f}%), increasing chunk size to {adaptive_chunk_size}"
                    )
                
                # Yield the chunk data and whether this is the final chunk
                logger.info(
                    f"Yielding chunk with {rows_in_current_chunk} records "
                    f"({total_rows_yielded}/{total_rows_to_process} processed)"
                )
                yield chunk_data, is_final_chunk
                
                # Force garbage collection to free memory between chunks
                # Also gives opportunity for checkpointing
                import gc
                gc.collect()
            
            logger.info(f"Completed streaming extraction with {processed_rows_count} processed rows.")
        
        except Exception as e:
            error_msg = f"Failed during streaming extraction: {str(e)}"
            logger.error(error_msg)
            raise DataExtractionError(error_msg) from e
    
    def _get_header_values_directly(self, sheet: Worksheet, header_row_index: int, header_map: Dict[int, str]) -> List[str]:
        """
        Helper method to get header values directly from a specific row.
        """
        headers = []
        min_col = 1 # Assuming columns start at 1
        for col_idx in range(min_col, max(header_map.keys()) + 1):
            header_value = sheet.get_cell_value(header_row_index, col_idx)
            headers.append(str(header_value) if header_value is not None else "")
        return headers
    
    def _process_row(
        self,
        sheet: Worksheet,
        row_data: Dict[int, Any],
        excel_row_idx: int,
        header_map: Dict[int, str],
        merge_map: Dict[Tuple[int, int], Dict],
        include_empty: bool
    ) -> HierarchicalRecord:
        """
        Process a single row of data, handling hierarchy from merges.
        
        Args:
            sheet: Worksheet accessor for getting cell values
            row_data: Dictionary mapping column indices to values for the row
            excel_row_idx: Excel row index (1-based)
            header_map: Dictionary mapping column indices to header names
            merge_map: Dictionary mapping (row, col) to merge information
            include_empty: Whether to include empty cells
            
        Returns:
            HierarchicalRecord with processed row data
        """
        record = HierarchicalRecord(row_index=excel_row_idx)

        current_record_items = {} # Temp dict to build the record items

        # Iterate through columns defined by the header map
        for col_idx, col_name in sorted(header_map.items()):
            current_cell_coords = (excel_row_idx, col_idx)
            
            # --- Get Value --- 
            value = None
            merge_info = None # Reset merge_info for each item
            is_merged_origin = False
            if current_cell_coords in merge_map:
                origin_data = merge_map[current_cell_coords]
                origin = origin_data['origin']
                # If this cell IS the origin of the merge
                if origin == current_cell_coords:
                    is_merged_origin = True
                    value = origin_data['value']
                    # Calculate minimal merge info (optional, could be omitted for pure key-value)
                    excel_range = origin_data['range']
                    range_parts = excel_range.split(':')
                    if len(range_parts) == 2:
                         cell1 = CellPosition.from_excel_notation(range_parts[0])
                         cell2 = CellPosition.from_excel_notation(range_parts[1])
                         row_span = cell2.row - cell1.row + 1
                         col_span = cell2.column - cell1.column + 1
                         merge_type = "block"
                         if row_span > 1 and col_span == 1: merge_type = "vertical"
                         elif row_span == 1 and col_span > 1: merge_type = "horizontal"
                         merge_info = MergeInfo(merge_type=merge_type, span=(row_span, col_span), range=excel_range, origin=origin)
                else:
                    # If part of merge but NOT origin, SKIP (value comes from origin)
                    continue 
            else:
                # Not part of any merge, get value directly from row data
                value = row_data.get(col_idx)
                 
            # Skip if value is None and we're not including empty cells (allow empty merge origins)
            if value is None and not include_empty and not is_merged_origin:
                continue

            # --- Position --- 
            position = CellPosition(row=excel_row_idx, column=col_idx)
            
            # --- Create Item (No Sub-items) --- 
            item = HierarchicalDataItem(
                key=str(col_name), # Key is the header name
                value=value,       # Value is the cell content (or merge origin content)
                position=position,
                merge_info=merge_info, # Store merge info if it was an origin
                sub_items=[]         # Explicitly empty sub-items
            )
            current_record_items[col_name] = item # Store with header name as key
            
        # Add collected items to the record
        for item in current_record_items.values():
            record.add_item(item)

        return record

    def extract_hierarchical_data(
        self,
        reader: ExcelReader,
        sheet_structure: SheetStructure,
        data_start_row: int,
        chunk_size: int = 1000,
        include_empty: bool = False
    ) -> HierarchicalData:
        """
        Extract hierarchical data with full context of sheet structure.
        
        Args:
            reader: ExcelReader instance
            sheet_structure: SheetStructure with sheet information
            data_start_row: Row number where data starts (including header)
            chunk_size: Number of rows to process at once
            include_empty: Whether to include empty cells
            
        Returns:
            HierarchicalData instance with extracted data
            
        Raises:
            HierarchicalDataError: If hierarchical data extraction fails
        """
        try:
            logger.info(
                f"Extracting hierarchical data from sheet: {sheet_structure.name} "
                f"starting at row {data_start_row}"
            )
            
            # Get the sheet
            sheet = reader.get_sheet(sheet_structure.name)
            
            # Extract data
            return self.extract_data(
                sheet=sheet,
                merge_map=sheet_structure.merge_map,
                data_start_row=data_start_row,
                chunk_size=chunk_size,
                include_empty=include_empty
            )
        except Exception as e:
            error_msg = f"Failed to extract hierarchical data: {str(e)}"
            logger.error(error_msg)
            raise HierarchicalDataError(
                error_msg,
                sheet_name=sheet_structure.name
            ) from e

    def _iterate_rows_chunked(self, sheet: Worksheet, start_row: int, end_row: int, chunk_size: int) -> Generator[Dict[int, Dict[int, Any]], None, None]:
        """
        Helper method to iterate through sheet rows in chunks using a generator.
        """
        min_row, max_row, min_col, max_col = sheet.get_dimensions()
        
        current_chunk = {}
        row_count_in_chunk = 0
        for row_idx in range(start_row, end_row + 1):
            # Ensure we don't try to read beyond actual sheet bounds if end_row was estimated high
            if row_idx > max_row:
                break 
                
            row_data = {}
            for col_idx in range(min_col, max_col + 1):
                row_data[col_idx] = sheet.get_cell_value(row_idx, col_idx)
            
            current_chunk[row_idx] = row_data
            row_count_in_chunk += 1
            
            if row_count_in_chunk == chunk_size:
                yield current_chunk
                current_chunk = {}
                row_count_in_chunk = 0
        
        # Yield any remaining rows in the last chunk
        if current_chunk:
            yield current_chunk