"""
Domain models for Excel structural elements.
Provides type-safe interfaces for cell ranges, merged regions, and sheet structure.
"""

from dataclasses import dataclass
from enum import Enum, auto
from typing import Dict, List, Optional, Tuple, Union


class CellDataType(Enum):
    """Enumeration of cell data types."""
    
    STRING = auto()
    NUMBER = auto()
    BOOLEAN = auto()
    DATE = auto()
    FORMULA = auto()
    ERROR = auto()
    EMPTY = auto()


@dataclass
class CellPosition:
    """Represents a position in an Excel sheet."""
    
    row: int
    column: int
    
    def to_tuple(self) -> Tuple[int, int]:
        """Convert to a (row, column) tuple."""
        return (self.row, self.column)
    
    def to_excel_notation(self) -> str:
        """Convert to Excel notation (e.g., 'A1')."""
        from openpyxl.utils import get_column_letter
        return f"{get_column_letter(self.column)}{self.row}"
    
    @classmethod
    def from_excel_notation(cls, notation: str) -> "CellPosition":
        """Create from Excel notation (e.g., 'A1')."""
        from openpyxl.utils import column_index_from_string
        # Extract column letters and row number
        import re
        match = re.match(r"([A-Za-z]+)([0-9]+)", notation)
        if not match:
            raise ValueError(f"Invalid Excel notation: {notation}")
        col_str, row_str = match.groups()
        return cls(
            row=int(row_str),
            column=column_index_from_string(col_str)
        )


@dataclass
class CellRange:
    """Represents a range of cells in an Excel sheet."""
    
    start: CellPosition
    end: CellPosition
    
    @property
    def width(self) -> int:
        """Get the width of the range (number of columns)."""
        return self.end.column - self.start.column + 1
    
    @property
    def height(self) -> int:
        """Get the height of the range (number of rows)."""
        return self.end.row - self.start.row + 1
    
    @property
    def size(self) -> Tuple[int, int]:
        """Get the size of the range as (height, width)."""
        return (self.height, self.width)
    
    def to_excel_notation(self) -> str:
        """Convert to Excel notation (e.g., 'A1:B2')."""
        return f"{self.start.to_excel_notation()}:{self.end.to_excel_notation()}"
    
    @classmethod
    def from_excel_notation(cls, notation: str) -> "CellRange":
        """Create from Excel notation (e.g., 'A1:B2')."""
        parts = notation.split(":")
        if len(parts) != 2:
            raise ValueError(f"Invalid range notation: {notation}")
        return cls(
            start=CellPosition.from_excel_notation(parts[0]),
            end=CellPosition.from_excel_notation(parts[1])
        )
    
    def contains(self, position: CellPosition) -> bool:
        """Check if this range contains the given position."""
        return (
            self.start.row <= position.row <= self.end.row and
            self.start.column <= position.column <= self.end.column
        )
    
    def iterate_positions(self) -> List[CellPosition]:
        """Iterate over all cell positions in this range."""
        positions = []
        for row in range(self.start.row, self.end.row + 1):
            for col in range(self.start.column, self.end.column + 1):
                positions.append(CellPosition(row=row, column=col))
        return positions


@dataclass
class MergedCell:
    """Represents a merged cell region in an Excel sheet."""
    
    range: CellRange
    value: Optional[object] = None
    
    @property
    def origin(self) -> CellPosition:
        """Get the top-left cell position (origin) of this merged cell."""
        return self.range.start
    
    @property
    def width(self) -> int:
        """Get the width of the merged cell (number of columns)."""
        return self.range.width
    
    @property
    def height(self) -> int:
        """Get the height of the merged cell (number of rows)."""
        return self.range.height
    
    @property
    def is_horizontal(self) -> bool:
        """Check if this is a horizontal merge (spans multiple columns, one row)."""
        return self.width > 1 and self.height == 1
    
    @property
    def is_vertical(self) -> bool:
        """Check if this is a vertical merge (spans multiple rows, one column)."""
        return self.height > 1 and self.width == 1
    
    @property
    def is_block(self) -> bool:
        """Check if this is a block merge (spans multiple rows and columns)."""
        return self.width > 1 and self.height > 1


@dataclass
class SheetDimensions:
    """Represents the dimensions of an Excel sheet."""
    
    min_row: int
    max_row: int
    min_column: int
    max_column: int
    
    @property
    def width(self) -> int:
        """Get the width of the sheet (number of columns)."""
        return self.max_column - self.min_column + 1
    
    @property
    def height(self) -> int:
        """Get the height of the sheet (number of rows)."""
        return self.max_row - self.min_row + 1
    
    @property
    def size(self) -> Tuple[int, int]:
        """Get the size of the sheet as (height, width)."""
        return (self.height, self.width)
    
    def to_cell_range(self) -> CellRange:
        """Convert to a CellRange representing the entire sheet."""
        return CellRange(
            start=CellPosition(row=self.min_row, column=self.min_column),
            end=CellPosition(row=self.max_row, column=self.max_column)
        )


@dataclass
class SheetStructure:
    """
    Represents the structure of an Excel sheet.
    Contains information about dimensions, merged cells, etc.
    """
    
    name: str
    dimensions: SheetDimensions
    merged_cells: List[MergedCell] = None
    merge_map: Dict[Tuple[int, int], Dict] = None
    
    def __post_init__(self) -> None:
        """Initialize default values for merged_cells and merge_map."""
        if self.merged_cells is None:
            self.merged_cells = []
        if self.merge_map is None:
            self.merge_map = {}
    
    @property
    def has_merged_cells(self) -> bool:
        """Check if this sheet has any merged cells."""
        return len(self.merged_cells) > 0