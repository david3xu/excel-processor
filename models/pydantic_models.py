"""
Pydantic models for Excel data structures.
Provides enhanced validation and serialization for Excel data.
"""

from enum import Enum, auto
from typing import Dict, List, Optional, Tuple, Union, Any

from pydantic import BaseModel, Field, validator


class CellDataType(Enum):
    """Enumeration of cell data types."""
    
    STRING = auto()
    NUMBER = auto()
    BOOLEAN = auto()
    DATE = auto()
    FORMULA = auto()
    ERROR = auto()
    EMPTY = auto()


class CellPosition(BaseModel):
    """Represents a position in an Excel sheet."""
    
    row: int = Field(..., ge=1, description="Row number (1-indexed)")
    column: int = Field(..., ge=1, description="Column number (1-indexed)")
    
    def to_tuple(self) -> Tuple[int, int]:
        """Convert to a (row, column) tuple."""
        return (self.row, self.column)
    
    def to_excel_notation(self) -> str:
        """Convert to Excel notation (e.g., 'A1')."""
        from openpyxl.utils import get_column_letter
        return f"{get_column_letter(self.column)}{self.row}"
    
    @classmethod
    def from_excel_notation(cls, notation: str) -> "CellPosition":
        """Create from Excel notation (e.g., 'A1')."""
        from openpyxl.utils import column_index_from_string
        # Extract column letters and row number
        import re
        match = re.match(r"([A-Za-z]+)([0-9]+)", notation)
        if not match:
            raise ValueError(f"Invalid Excel notation: {notation}")
        col_str, row_str = match.groups()
        return cls(
            row=int(row_str),
            column=column_index_from_string(col_str)
        )
    
    class Config:
        """Pydantic configuration for CellPosition."""
        extra = "forbid"


class CellRange(BaseModel):
    """Represents a range of cells in an Excel sheet."""
    
    start: CellPosition
    end: CellPosition
    
    @property
    def width(self) -> int:
        """Get the width of the range (number of columns)."""
        return self.end.column - self.start.column + 1
    
    @property
    def height(self) -> int:
        """Get the height of the range (number of rows)."""
        return self.end.row - self.start.row + 1
    
    @property
    def size(self) -> Tuple[int, int]:
        """Get the size of the range as (height, width)."""
        return (self.height, self.width)
    
    def to_excel_notation(self) -> str:
        """Convert to Excel notation (e.g., 'A1:B2')."""
        return f"{self.start.to_excel_notation()}:{self.end.to_excel_notation()}"
    
    @classmethod
    def from_excel_notation(cls, notation: str) -> "CellRange":
        """Create from Excel notation (e.g., 'A1:B2')."""
        parts = notation.split(":")
        if len(parts) != 2:
            raise ValueError(f"Invalid range notation: {notation}")
        return cls(
            start=CellPosition.from_excel_notation(parts[0]),
            end=CellPosition.from_excel_notation(parts[1])
        )
    
    def contains(self, position: CellPosition) -> bool:
        """Check if this range contains the given position."""
        return (
            self.start.row <= position.row <= self.end.row and
            self.start.column <= position.column <= self.end.column
        )
    
    def iterate_positions(self) -> List[CellPosition]:
        """Iterate over all cell positions in this range."""
        positions = []
        for row in range(self.start.row, self.end.row + 1):
            for col in range(self.start.column, self.end.column + 1):
                positions.append(CellPosition(row=row, column=col))
        return positions
    
    @validator('end')
    def validate_range(cls, v, values):
        """Validate that the end position is after the start position."""
        if 'start' in values:
            start = values['start']
            if v.row < start.row or v.column < start.column:
                raise ValueError(
                    f"End position ({v.row}, {v.column}) must be after start position "
                    f"({start.row}, {start.column})"
                )
        return v
    
    class Config:
        """Pydantic configuration for CellRange."""
        extra = "forbid"


class MergedCell(BaseModel):
    """Represents a merged cell region in an Excel sheet."""
    
    range: CellRange
    value: Optional[Any] = None
    
    @property
    def origin(self) -> CellPosition:
        """Get the top-left cell position (origin) of this merged cell."""
        return self.range.start
    
    @property
    def width(self) -> int:
        """Get the width of the merged cell (number of columns)."""
        return self.range.width
    
    @property
    def height(self) -> int:
        """Get the height of the merged cell (number of rows)."""
        return self.range.height
    
    @property
    def is_horizontal(self) -> bool:
        """Check if this is a horizontal merge (spans multiple columns, one row)."""
        return self.width > 1 and self.height == 1
    
    @property
    def is_vertical(self) -> bool:
        """Check if this is a vertical merge (spans multiple rows, one column)."""
        return self.height > 1 and self.width == 1
    
    @property
    def is_block(self) -> bool:
        """Check if this is a block merge (spans multiple rows and columns)."""
        return self.width > 1 and self.height > 1
    
    class Config:
        """Pydantic configuration for MergedCell."""
        extra = "forbid"
        arbitrary_types_allowed = True


class SheetDimensions(BaseModel):
    """Represents the dimensions of an Excel sheet."""
    
    min_row: int = Field(..., ge=1, description="Minimum row number")
    max_row: int = Field(..., ge=1, description="Maximum row number")
    min_column: int = Field(..., ge=1, description="Minimum column number")
    max_column: int = Field(..., ge=1, description="Maximum column number")
    
    @property
    def width(self) -> int:
        """Get the width of the sheet (number of columns)."""
        return self.max_column - self.min_column + 1
    
    @property
    def height(self) -> int:
        """Get the height of the sheet (number of rows)."""
        return self.max_row - self.min_row + 1
    
    @property
    def size(self) -> Tuple[int, int]:
        """Get the size of the sheet as (height, width)."""
        return (self.height, self.width)
    
    def to_cell_range(self) -> CellRange:
        """Convert to a CellRange representing the entire sheet."""
        return CellRange(
            start=CellPosition(row=self.min_row, column=self.min_column),
            end=CellPosition(row=self.max_row, column=self.max_column)
        )
    
    @validator('max_row')
    def validate_max_row(cls, v, values):
        """Validate max_row is greater than or equal to min_row."""
        if 'min_row' in values and v < values['min_row']:
            raise ValueError(f"max_row ({v}) must be >= min_row ({values['min_row']})")
        return v
    
    @validator('max_column')
    def validate_max_column(cls, v, values):
        """Validate max_column is greater than or equal to min_column."""
        if 'min_column' in values and v < values['min_column']:
            raise ValueError(f"max_column ({v}) must be >= min_column ({values['min_column']})")
        return v
    
    class Config:
        """Pydantic configuration for SheetDimensions."""
        extra = "forbid"


class SheetStructure(BaseModel):
    """
    Represents the structure of an Excel sheet.
    Contains information about dimensions, merged cells, etc.
    """
    
    name: str = Field(..., description="Sheet name")
    dimensions: SheetDimensions
    merged_cells: List[MergedCell] = Field(default_factory=list, description="List of merged cell regions")
    merge_map: Dict[Tuple[int, int], Dict] = Field(default_factory=dict, description="Map of merged cells by position")
    
    @property
    def has_merged_cells(self) -> bool:
        """Check if this sheet has any merged cells."""
        return len(self.merged_cells) > 0
    
    class Config:
        """Pydantic configuration for SheetStructure."""
        extra = "forbid"
        arbitrary_types_allowed = True  # For complex types like Tuple as dict keys


class ExcelCell(BaseModel):
    """
    Represents a cell in an Excel sheet with its value and metadata.
    """
    value: Any = None 
    row: int = Field(..., ge=1, description="Row number (1-indexed)")
    column: int = Field(..., ge=1, description="Column number (1-indexed)")
    column_name: Optional[str] = Field(None, description="Column name/letter")
    data_type: Optional[CellDataType] = Field(None, description="Data type of the cell value")
    
    class Config:
        """Pydantic configuration for ExcelCell."""
        extra = "forbid"
        arbitrary_types_allowed = True


class ExcelRow(BaseModel):
    """
    Represents a row in an Excel sheet with its cells.
    """
    cells: Dict[str, ExcelCell] = Field(default_factory=dict, description="Map of column names to cells")
    row_number: int = Field(..., ge=1, description="Row number (1-indexed)")
    
    def to_dict(self) -> Dict[str, Any]:
        """Convert to a simple dictionary of column names to values."""
        return {k: v.value for k, v in self.cells.items()}
    
    class Config:
        """Pydantic configuration for ExcelRow."""
        extra = "forbid"


class ExcelSheet(BaseModel):
    """
    Represents a sheet in an Excel file with its rows and metadata.
    """
    name: str = Field(..., description="Sheet name")
    rows: List[ExcelRow] = Field(default_factory=list, description="Rows in the sheet")
    headers: List[str] = Field(default_factory=list, description="Column headers")
    structure: Optional[SheetStructure] = Field(None, description="Structural information about the sheet")
    
    class Config:
        """Pydantic configuration for ExcelSheet."""
        extra = "forbid" 