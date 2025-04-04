import pandas as pd
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from datetime import datetime

def create_test_excel(filename="knowledge_graph_test_data.xlsx"):
    """Create a test Excel file with merged cells and hierarchical data for testing the Excel Processor."""
    
    # Create a new Excel writer
    writer = pd.ExcelWriter(filename, engine='openpyxl')
    
    # ------------------------------------------------------------------------
    # Sheet 1: Equipment Failure Events
    # ------------------------------------------------------------------------
    # Create the data frame for Sheet 1
    data1 = [
        ["MECHANICAL", "Pump System", "Bearing Failure", "Critical", "2024-03-15"],
        ["", "", "Seal Leakage", "Medium", "2024-02-20"],
        ["", "Conveyor System", "Belt Misalignment", "Low", "2024-03-03"],
        ["", "", "Drive Failure", "High", "2024-01-25"],
        ["ELECTRICAL", "Motor Control", "Overload Trip", "Medium", "2024-03-10"],
        ["", "", "Phase Imbalance", "Medium", "2024-02-12"],
        ["", "Power Supply", "Voltage Fluctuation", "Low", "2024-03-18"]
    ]
    
    df1 = pd.DataFrame(data1, columns=["Event Category", "Equipment Class", "Failure Event", "Impact Level", "Date"])
    
    # Convert dates to datetime objects
    df1["Date"] = pd.to_datetime(df1["Date"])
    
    # Write the dataframe to Sheet 1
    df1.to_excel(writer, sheet_name="Equipment Failure Events", index=False)
    
    # ------------------------------------------------------------------------
    # Sheet 2: Root Cause Analysis
    # ------------------------------------------------------------------------
    # Create the data frame for Sheet 2
    data2 = [
        ["EV-2024-001\n(Bearing Failure)", "Maintenance", "Improper Lubrication", "- Lubrication Schedule Missed (2024-01-10)\n- Vibration Monitoring Gap (2023-12-15)\n- Training Deficiency"],
        ["EV-2024-002\n(Seal Leakage)", "Operational", "Process Deviation", "- Parameter Setting Error\n- Operator Handover Communication"],
        ["EV-2024-003\n(Belt Misalign)", "Design", "Material Selection", "- Environmental Exposure\n- Load Calculation Error"]
    ]
    
    df2 = pd.DataFrame(data2, columns=["Failure Event ID", "Root Cause Category", "Primary Cause", "Contributing Factors"])
    
    # Write the dataframe to Sheet 2
    df2.to_excel(writer, sheet_name="Root Cause Analysis", index=False, startrow=1)
    
    # ------------------------------------------------------------------------
    # Sheet 3: Knowledge Graph Relationships
    # ------------------------------------------------------------------------
    # Create the data frame for Sheet 3
    data3 = [
        ["Bearing Failure", "causedBy", "Improper Lubrication", 95],
        ["Improper Lubrication", "contributedBy", "Lubrication Schedule Missed", 90],
        ["Improper Lubrication", "contributedBy", "Vibration Monitoring Gap", 85],
        ["Improper Lubrication", "contributedBy", "Training Deficiency", 80],
        ["Lubrication Schedule Missed", "affects", "Bearing Component", 95],
        ["Training Deficiency", "influences", "Maintenance Quality", 75],
        ["Training Deficiency", "influences", "Operational Compliance", 70]
    ]
    
    df3 = pd.DataFrame(data3, columns=["Source Entity", "Relationship", "Target Entity", "Confidence (%)"])
    
    # Write the dataframe to Sheet 3
    df3.to_excel(writer, sheet_name="Knowledge Graph Relationships", index=False)
    
    # ------------------------------------------------------------------------
    # Apply formatting and create merged cells
    # ------------------------------------------------------------------------
    workbook = writer.book
    
    # Format Sheet 1: Equipment Failure Events
    sheet1 = workbook["Equipment Failure Events"]
    
    # Apply merged cells for Event Category
    sheet1.merge_cells('A2:A5')  # MECHANICAL
    sheet1.merge_cells('A6:A8')  # ELECTRICAL
    
    # Apply merged cells for Equipment Class
    sheet1.merge_cells('B2:B3')  # Pump System
    sheet1.merge_cells('B4:B5')  # Conveyor System
    sheet1.merge_cells('B6:B7')  # Motor Control
    
    # Format Sheet 2: Root Cause Analysis
    sheet2 = workbook["Root Cause Analysis"]
    
    # Add metadata header
    sheet2.merge_cells('A1:D1')
    sheet2["A1"] = "Plant: Manufacturing Line A                      Analysis Period: Q1 2024        Analyst: J. Smith"
    sheet2["A1"].alignment = Alignment(horizontal='left', wrap_text=True)
    sheet2["A1"].font = Font(bold=True)
    sheet2["A1"].fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    
    # Format Sheet 3: Knowledge Graph Relationships
    sheet3 = workbook["Knowledge Graph Relationships"]
    
    # Apply merged cells for Source Entity
    sheet3.merge_cells('A2:A2')  # Bearing Failure
    sheet3.merge_cells('A3:A6')  # Improper Lubrication
    sheet3.merge_cells('A7:A7')  # Lubrication Schedule Missed
    sheet3.merge_cells('A8:A9')  # Training Deficiency
    
    # Format headers in all sheets
    for sheet in [sheet1, sheet2, sheet3]:
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
    
    # Add borders
    thin_border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    
    for sheet in [sheet1, sheet2, sheet3]:
        for row in sheet.iter_rows():
            for cell in row:
                cell.border = thin_border
    
    # Adjust column widths
    for sheet in [sheet1, sheet2, sheet3]:
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column_letter].width = adjusted_width
    
    # Save the workbook
    writer.close()
    
    print(f"Test Excel file created: {filename}")
    return filename

if __name__ == "__main__":
    create_test_excel()