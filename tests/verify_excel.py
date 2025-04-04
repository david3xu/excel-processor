#!/usr/bin/env python
"""
Script to verify that the generated Excel file can be correctly read.
"""

import sys
import openpyxl
from openpyxl.utils import get_column_letter

def verify_excel(excel_file):
    """
    Try to read an Excel file using openpyxl directly to verify it's valid.
    """
    print(f"Attempting to verify Excel file: {excel_file}")
    
    try:
        # Try to load the workbook
        print("Loading workbook...")
        wb = openpyxl.load_workbook(excel_file)
        
        # Print basic workbook info
        print(f"Workbook loaded successfully!")
        print(f"Sheets: {wb.sheetnames}")
        
        # Try to read from each sheet
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            print(f"\nSheet: {sheet_name}")
            print(f"Dimensions: {sheet.dimensions}")
            print(f"Max row: {sheet.max_row}, Max column: {sheet.max_column}")
            
            # Print sample data (first 5 rows)
            print("\nSample data:")
            for row in range(1, min(6, sheet.max_row + 1)):
                row_data = []
                for col in range(1, min(6, sheet.max_column + 1)):
                    cell = sheet.cell(row=row, column=col)
                    row_data.append(str(cell.value))
                print(f"Row {row}: {' | '.join(row_data)}")
            
            # Check for merged cells
            if sheet.merged_cells:
                print(f"\nMerged cells: {len(sheet.merged_cells.ranges)}")
                for merged_range in sheet.merged_cells.ranges:
                    print(f"  {merged_range} = {sheet.cell(merged_range.min_row, merged_range.min_col).value}")
        
        print("\nExcel file validated successfully!")
        return True
        
    except Exception as e:
        print(f"Error verifying Excel file: {type(e).__name__}: {str(e)}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python verify_excel.py <excel_file>")
        sys.exit(1)
    
    excel_file = sys.argv[1]
    if verify_excel(excel_file):
        sys.exit(0)
    else:
        sys.exit(1) 