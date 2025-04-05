import pandas as pd
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from datetime import datetime
import os

def create_test_excel(filename="knowledge_graph_test_data.xlsx"):
    """Create a test Excel file with merged cells and hierarchical data for testing the Excel Processor."""
    
    # Define the output directory
    output_dir = os.path.join("data", "input")
    
    # Create the directory if it doesn't exist
    os.makedirs(output_dir, exist_ok=True)
    
    # Create the full file path
    file_path = os.path.join(output_dir, filename)
    
    # Create a new workbook directly with openpyxl
    workbook = openpyxl.Workbook()
    
    # ------------------------------------------------------------------------
    # Sheet 1: Equipment Failure Events
    # ------------------------------------------------------------------------
    # Rename the default sheet
    sheet1 = workbook.active
    sheet1.title = "Equipment Failure Events"
    
    # Define headers
    headers = ["Event Category", "Equipment Class", "Failure Event", "Impact Level", "Date"]
    
    # Write headers
    for col_idx, header in enumerate(headers, start=1):
        sheet1.cell(row=1, column=col_idx, value=header)
    
    # Define data
    data1 = [
        ["MECHANICAL", "Pump System", "Bearing Failure", "Critical", "2024-03-15"],
        ["", "", "Seal Leakage", "Medium", "2024-02-20"],
        ["", "Conveyor System", "Belt Misalignment", "Low", "2024-03-03"],
        ["", "", "Drive Failure", "High", "2024-01-25"],
        ["ELECTRICAL", "Motor Control", "Overload Trip", "Medium", "2024-03-10"],
        ["", "", "Phase Imbalance", "Medium", "2024-02-12"],
        ["", "Power Supply", "Voltage Fluctuation", "Low", "2024-03-18"]
    ]
    
    # Write data
    for row_idx, row_data in enumerate(data1, start=2):
        for col_idx, cell_value in enumerate(row_data, start=1):
            # Convert dates to datetime objects
            if col_idx == 5 and cell_value:
                cell_value = datetime.strptime(cell_value, "%Y-%m-%d")
            
            sheet1.cell(row=row_idx, column=col_idx, value=cell_value)
    
    # ------------------------------------------------------------------------
    # Sheet 2: Root Cause Analysis
    # ------------------------------------------------------------------------
    sheet2 = workbook.create_sheet(title="Root Cause Analysis")
    
    # Define data
    headers2 = ["Failure Event ID", "Root Cause Category", "Primary Cause", "Contributing Factors"]
    
    # Add metadata header
    sheet2.cell(row=1, column=1, value="Plant: Manufacturing Line A                      Analysis Period: Q1 2024        Analyst: J. Smith")
    
    # Write headers
    for col_idx, header in enumerate(headers2, start=1):
        sheet2.cell(row=2, column=col_idx, value=header)
    
    # Define data
    data2 = [
        ["EV-2024-001\n(Bearing Failure)", "Maintenance", "Improper Lubrication", "- Lubrication Schedule Missed (2024-01-10)\n- Vibration Monitoring Gap (2023-12-15)\n- Training Deficiency"],
        ["EV-2024-002\n(Seal Leakage)", "Operational", "Process Deviation", "- Parameter Setting Error\n- Operator Handover Communication"],
        ["EV-2024-003\n(Belt Misalign)", "Design", "Material Selection", "- Environmental Exposure\n- Load Calculation Error"]
    ]
    
    # Write data
    for row_idx, row_data in enumerate(data2, start=3):
        for col_idx, cell_value in enumerate(row_data, start=1):
            cell = sheet2.cell(row=row_idx, column=col_idx, value=cell_value)
            # Set text wrapping for multi-line text
            cell.alignment = Alignment(wrap_text=True)
    
    # ------------------------------------------------------------------------
    # Sheet 3: Knowledge Graph Relationships
    # ------------------------------------------------------------------------
    sheet3 = workbook.create_sheet(title="Knowledge Graph Relationships")
    
    # Define headers
    headers3 = ["Source Entity", "Relationship", "Target Entity", "Confidence (%)"]
    
    # Write headers
    for col_idx, header in enumerate(headers3, start=1):
        sheet3.cell(row=1, column=col_idx, value=header)
    
    # Define data
    data3 = [
        ["Bearing Failure", "causedBy", "Improper Lubrication", 95],
        ["Improper Lubrication", "contributedBy", "Lubrication Schedule Missed", 90],
        ["Improper Lubrication", "contributedBy", "Vibration Monitoring Gap", 85],
        ["Improper Lubrication", "contributedBy", "Training Deficiency", 80],
        ["Lubrication Schedule Missed", "affects", "Bearing Component", 95],
        ["Training Deficiency", "influences", "Maintenance Quality", 75],
        ["Training Deficiency", "influences", "Operational Compliance", 70]
    ]
    
    # Write data
    for row_idx, row_data in enumerate(data3, start=2):
        for col_idx, cell_value in enumerate(row_data, start=1):
            sheet3.cell(row=row_idx, column=col_idx, value=cell_value)
    
    # ------------------------------------------------------------------------
    # Apply formatting and create merged cells
    # ------------------------------------------------------------------------
    
    # Apply merged cells for Sheet 1: Equipment Failure Events
    sheet1.merge_cells('A2:A5')  # MECHANICAL
    sheet1.merge_cells('A6:A8')  # ELECTRICAL
    sheet1.merge_cells('B2:B3')  # Pump System
    sheet1.merge_cells('B4:B5')  # Conveyor System
    sheet1.merge_cells('B6:B7')  # Motor Control
    
    # Apply merged cells for Sheet 2: Root Cause Analysis
    sheet2.merge_cells('A1:D1')
    sheet2.cell(row=1, column=1).alignment = Alignment(horizontal='left', wrap_text=True)
    sheet2.cell(row=1, column=1).font = Font(bold=True)
    sheet2.cell(row=1, column=1).fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    
    # Apply merged cells for Sheet 3: Knowledge Graph Relationships
    sheet3.merge_cells('A2:A2')  # Bearing Failure
    sheet3.merge_cells('A3:A6')  # Improper Lubrication
    sheet3.merge_cells('A7:A7')  # Lubrication Schedule Missed
    sheet3.merge_cells('A8:A9')  # Training Deficiency
    
    # Format headers in all sheets
    for sheet in [sheet1, sheet2, sheet3]:
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
    
    # Add borders
    thin_border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    
    for sheet in [sheet1, sheet2, sheet3]:
        for row in sheet.iter_rows():
            for cell in row:
                cell.border = thin_border
    
    # ------------------------------------------------------------------------
    # Adjust column widths for all sheets
    # ------------------------------------------------------------------------
    for sheet in workbook.worksheets:
        for column_idx in range(1, sheet.max_column + 1):
            max_length = 0
            column_letter = openpyxl.utils.get_column_letter(column_idx)
            
            # Get all cells in the column
            for row_idx in range(1, sheet.max_row + 1):
                cell = sheet.cell(row=row_idx, column=column_idx)
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass
            
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column_letter].width = adjusted_width
    
    # Save the workbook directly with openpyxl
    workbook.save(file_path)
    
    print(f"Test Excel file created: {file_path}")
    return file_path

def create_complex_headers_excel(filename="complex_headers_test.xlsx"):
    """
    Create a complex Excel file with multi-level headers, merged cells, and various data types
    to thoroughly test the header preservation functionality.
    """
    # Define the output directory
    output_dir = os.path.join("data", "input")
    
    # Create the directory if it doesn't exist
    os.makedirs(output_dir, exist_ok=True)
    
    # Create the full file path
    file_path = os.path.join(output_dir, filename)
    
    # Create a new workbook
    workbook = openpyxl.Workbook()
    
    # ------------------------------------------------------------------------
    # Sheet 1: Multi-level Headers with Merged Cells
    # ------------------------------------------------------------------------
    sheet1 = workbook.active
    sheet1.title = "Multi-level Headers"
    
    # Add title as metadata
    sheet1.cell(row=1, column=1, value="Plant Operations Data Report")
    sheet1.merge_cells('A1:H1')
    sheet1.cell(row=1, column=1).font = Font(bold=True, size=14)
    sheet1.cell(row=1, column=1).alignment = Alignment(horizontal='center')
    sheet1.cell(row=1, column=1).fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    
    # Add date info
    sheet1.cell(row=2, column=1, value="Report Date:")
    sheet1.cell(row=2, column=2, value=datetime.now().date())
    sheet1.cell(row=2, column=2).number_format = 'YYYY-MM-DD'
    
    # Multi-level headers (3 levels)
    # Level 1 - Main Categories
    sheet1.cell(row=4, column=1, value="Equipment ID")
    sheet1.merge_cells('A4:A6')  # Merge vertically for all 3 header levels
    
    sheet1.cell(row=4, column=2, value="Production Metrics")
    sheet1.merge_cells('B4:D4')  # Merge horizontally across 3 columns
    
    sheet1.cell(row=4, column=5, value="Quality Metrics")
    sheet1.merge_cells('E4:F4')  # Merge horizontally across 2 columns
    
    sheet1.cell(row=4, column=7, value="Maintenance")
    sheet1.merge_cells('G4:H4')  # Merge horizontally across 2 columns
    
    # Level 2 - Sub-categories
    sheet1.cell(row=5, column=2, value="Daily Output")
    sheet1.merge_cells('B5:C5')  # Merge horizontally across 2 columns
    
    sheet1.cell(row=5, column=4, value="Efficiency")
    sheet1.cell(row=5, column=5, value="Defect Rate")
    sheet1.cell(row=5, column=6, value="Quality Score")
    sheet1.cell(row=5, column=7, value="Downtime")
    sheet1.cell(row=5, column=8, value="Maintenance Cost")
    
    # Level 3 - Detail headers
    sheet1.cell(row=6, column=2, value="Units")
    sheet1.cell(row=6, column=3, value="Weight (kg)")
    sheet1.cell(row=6, column=4, value="%")
    sheet1.cell(row=6, column=5, value="ppm")
    sheet1.cell(row=6, column=6, value="1-5")
    sheet1.cell(row=6, column=7, value="Hours")
    sheet1.cell(row=6, column=8, value="$")
    
    # Format all header cells
    for row in range(4, 7):
        for col in range(1, 9):
            cell = sheet1.cell(row=row, column=col)
            if cell.value:  # Only style cells with values
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")
    
    # Sample data
    data1 = [
        ["EQ-001", 547, 2735, 92.5, 350, 4.2, 3.5, 1250.75],
        ["EQ-002", 423, 2115, 88.7, 420, 3.8, 5.2, 1875.50],
        ["EQ-003", 612, 3060, 95.2, 175, 4.7, 1.8, 920.25],
        ["EQ-004", 380, 1900, 86.3, 510, 3.3, 8.7, 2340.80],
        ["EQ-005", None, None, 90.1, 280, 4.0, 4.5, 1500.00],  # Some empty cells
        ["EQ-006", 530, 2650, None, None, None, 2.3, 1150.40],  # More empty cells
    ]
    
    # Write data
    for row_idx, row_data in enumerate(data1, start=7):  # Start after the header rows
        for col_idx, value in enumerate(row_data, start=1):
            sheet1.cell(row=row_idx, column=col_idx, value=value)
    
    # ------------------------------------------------------------------------
    # Sheet 2: Complex Mixed Headers with Different Data Types
    # ------------------------------------------------------------------------
    sheet2 = workbook.create_sheet(title="Mixed Data Types")
    
    # Simple metadata
    sheet2.cell(row=1, column=1, value="Mixed Data Types Example")
    sheet2.merge_cells('A1:G1')
    sheet2.cell(row=1, column=1).font = Font(bold=True)
    sheet2.cell(row=1, column=1).alignment = Alignment(horizontal='center')
    
    # Headers with different data types
    headers2 = [
        "Text Header",           # Text
        123,                     # Number
        datetime.now().date(),   # Date
        True,                    # Boolean
        "=SUM(B3:B10)",          # Formula
        0.25,                    # Percentage
        "#N/A"                   # Error value
    ]
    
    # Write headers
    for col_idx, header in enumerate(headers2, start=1):
        cell = sheet2.cell(row=2, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        
        # Format cells according to their data type
        if isinstance(header, datetime):
            cell.number_format = 'YYYY-MM-DD'
        elif isinstance(header, float) and col_idx == 6:
            cell.number_format = '0.00%'
            
    # Sample data
    data2 = [
        ["Sample 1", 42, datetime(2024, 1, 15), True, 123.45, 0.75, "Value"],
        ["Sample 2", 18, datetime(2024, 2, 20), False, 67.89, 0.30, "Test"],
        ["Sample 3", 93, datetime(2024, 3, 10), True, 456.78, 0.15, None],
        ["Sample 4", 27, datetime(2024, 4, 5), False, 234.56, 0.50, "Data"],
    ]
    
    # Write data
    for row_idx, row_data in enumerate(data2, start=3):
        for col_idx, value in enumerate(row_data, start=1):
            cell = sheet2.cell(row=row_idx, column=col_idx, value=value)
            if isinstance(value, datetime):
                cell.number_format = 'YYYY-MM-DD'
            elif isinstance(value, float) and col_idx == 6:
                cell.number_format = '0.00%'
    
    # ------------------------------------------------------------------------
    # Sheet 3: Multiple Header Rows with Irregular Structure
    # ------------------------------------------------------------------------
    sheet3 = workbook.create_sheet(title="Irregular Headers")
    
    # Create a complex header structure with gaps and irregular merged cells
    sheet3.cell(row=1, column=1, value="Region")
    sheet3.merge_cells('A1:A3')  # Merge first column vertically
    
    sheet3.cell(row=1, column=2, value="North America")
    sheet3.merge_cells('B1:D1')  # Merge horizontally
    
    sheet3.cell(row=1, column=5, value="Europe")
    sheet3.merge_cells('E1:G1')  # Merge horizontally
    
    # Second level headers
    sheet3.cell(row=2, column=2, value="USA")
    sheet3.merge_cells('B2:C2')  # Merge USA across two columns
    sheet3.cell(row=2, column=4, value="Canada")
    sheet3.cell(row=2, column=5, value="UK")
    sheet3.cell(row=2, column=6, value="Germany")
    sheet3.cell(row=2, column=7, value="France")
    
    # Third level headers
    sheet3.cell(row=3, column=2, value="Q1")
    sheet3.cell(row=3, column=3, value="Q2")
    sheet3.cell(row=3, column=4, value="Q1")
    sheet3.cell(row=3, column=5, value="Q1")
    sheet3.cell(row=3, column=6, value="Q1")
    sheet3.cell(row=3, column=7, value="Q1")
    
    # Format header cells
    for row in range(1, 4):
        for col in range(1, 8):
            cell = sheet3.cell(row=row, column=col)
            if cell.value:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Different colors for different header levels
                if row == 1:
                    cell.fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
                elif row == 2:
                    cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
                
                cell.font = Font(bold=True, color="FFFFFF")
    
    # Sample data - product sales by region
    products = ["Product A", "Product B", "Product C", "Product D"]
    data3 = [
        [p, 100 + i*10, 200 + i*15, 150 + i*5, 180 + i*12, 160 + i*8, 140 + i*7] 
        for i, p in enumerate(products)
    ]
    
    # Write data
    for row_idx, row_data in enumerate(data3, start=4):
        for col_idx, value in enumerate(row_data, start=1):
            sheet3.cell(row=row_idx, column=col_idx, value=value)
    
    # ------------------------------------------------------------------------
    # Sheet 4: Sparse Data with Empty Headers
    # ------------------------------------------------------------------------
    sheet4 = workbook.create_sheet(title="Sparse Data")
    
    # Create a sparse header structure with empty cells
    headers4 = ["", "January", "", "February", "", "March", ""]
    for col_idx, header in enumerate(headers4, start=1):
        cell = sheet4.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        if header:
            cell.fill = PatternFill(start_color="9BC2E6", end_color="9BC2E6", fill_type="solid")
    
    # Second header row with some empty cells
    subheaders = ["Metric", "Value", "Change", "Value", "Change", "Value", "Change"]
    for col_idx, header in enumerate(subheaders, start=1):
        cell = sheet4.cell(row=2, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    
    # Sparse data with empty cells
    metrics = ["Revenue", "Costs", "Profit", "Units Sold", "Returns"]
    sparse_data = [
        ["Revenue", 12500, "+5%", 13200, "+6%", 14100, "+7%"],
        ["Costs", 8200, "+3%", "", "", 8500, "+4%"],
        ["Profit", 4300, "+8%", 4700, "+9%", "", ""],
        ["Units Sold", "", "", 1050, "+12%", 1150, "+10%"],
        ["Returns", 120, "-2%", 115, "-4%", 105, "-8%"]
    ]
    
    # Write data
    for row_idx, row_data in enumerate(sparse_data, start=3):
        for col_idx, value in enumerate(row_data, start=1):
            sheet4.cell(row=row_idx, column=col_idx, value=value)
    
    # ------------------------------------------------------------------------
    # Adjust column widths for all sheets
    # ------------------------------------------------------------------------
    for sheet in workbook.worksheets:
        for column_idx in range(1, sheet.max_column + 1):
            max_length = 0
            column_letter = openpyxl.utils.get_column_letter(column_idx)
            
            # Get all cells in the column
            for row_idx in range(1, sheet.max_row + 1):
                cell = sheet.cell(row=row_idx, column=column_idx)
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass
            
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column_letter].width = adjusted_width
    
    # Save the workbook
    workbook.save(file_path)
    print(f"Complex headers test Excel file created: {file_path}")
    return file_path

if __name__ == "__main__":
    create_test_excel()
    create_complex_headers_excel()