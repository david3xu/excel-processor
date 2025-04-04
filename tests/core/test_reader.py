import unittest
from unittest.mock import MagicMock, patch
from pathlib import Path

# Import necessary modules from core.reader
from core.reader import ExcelReader
from excel_io.strategies.fallback_strategy import SheetNotFoundError
from utils.exceptions import FileNotFoundError, FileReadError, ExcelReadError # Import required exceptions

# Assuming openpyxl classes are needed for type hints/mocks
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


class TestExcelReader(unittest.TestCase):
    """Test suite for the ExcelReader class in core/reader.py."""

    def setUp(self):
        """Set up test fixtures, if any."""
        self.mock_file_path = "non_existent_file.xlsx"

    def tearDown(self):
        """Tear down test fixtures, if any."""
        pass

    @patch('core.reader.Path.exists')
    def test_init_file_not_found(self, mock_exists):
        print(f"--- Running: {self.__class__.__name__}.test_init_file_not_found ---")
        """Test __init__ raises FileNotFoundError if file does not exist."""
        # Arrange
        mock_exists.return_value = False

        # Act & Assert
        with self.assertRaises(FileNotFoundError) as cm:
            ExcelReader(self.mock_file_path)
        self.assertIn("Excel file not found", str(cm.exception))
        mock_exists.assert_called_once()

    @patch('core.reader.openpyxl.load_workbook')
    @patch('core.reader.Path.exists')
    def test_load_workbook_success(self, mock_exists, mock_load):
        print(f"--- Running: {self.__class__.__name__}.test_load_workbook_success ---")
        """Test load_workbook successfully loads and stores the workbook."""
        # Arrange
        mock_exists.return_value = True
        mock_workbook = MagicMock(spec=Workbook)
        mock_active_sheet = MagicMock(spec=Worksheet)
        mock_workbook.active = mock_active_sheet
        mock_workbook.sheetnames = ["Sheet1", "Sheet2"]
        mock_load.return_value = mock_workbook

        reader = ExcelReader(self.mock_file_path)

        # Act
        loaded_wb = reader.load_workbook(data_only=False)

        # Assert
        self.assertEqual(loaded_wb, mock_workbook)
        self.assertEqual(reader.workbook, mock_workbook)
        self.assertEqual(reader.active_sheet, mock_active_sheet)
        mock_load.assert_called_once_with(reader.excel_file, data_only=False)

    @patch('core.reader.openpyxl.load_workbook')
    @patch('core.reader.Path.exists')
    def test_load_workbook_failure(self, mock_exists, mock_load):
        print(f"--- Running: {self.__class__.__name__}.test_load_workbook_failure ---")
        """Test load_workbook raises FileReadError on openpyxl failure."""
        # Arrange
        mock_exists.return_value = True
        mock_load.side_effect = Exception("Openpyxl failed")

        reader = ExcelReader(self.mock_file_path)

        # Act & Assert
        with self.assertRaises(FileReadError) as cm:
            reader.load_workbook()
        self.assertIn("Failed to load workbook: Openpyxl failed", str(cm.exception))
        mock_load.assert_called_once()

    def test_get_sheet_names(self):
        print(f"--- Running: {self.__class__.__name__}.test_get_sheet_names ---")
        """Test get_sheet_names returns the list of names from the workbook."""
        # Arrange
        # Mock Path.exists to allow instantiation
        with patch('core.reader.Path.exists', return_value=True):
            reader = ExcelReader(self.mock_file_path)
            # Manually set the mocked workbook after instantiation
            mock_workbook = MagicMock(spec=Workbook)
            mock_workbook.sheetnames = ["DataSheet", "MetadataSheet"]
            reader.workbook = mock_workbook # Assign mocked workbook

            # Act
            sheet_names = reader.get_sheet_names()

            # Assert
            self.assertEqual(sheet_names, ["DataSheet", "MetadataSheet"])

    def test_get_sheet_names_workbook_not_loaded(self):
        print(f"--- Running: {self.__class__.__name__}.test_get_sheet_names_workbook_not_loaded ---")
        """Test get_sheet_names raises error if workbook is not loaded."""
        # Arrange
        with patch('core.reader.Path.exists', return_value=True):
            reader = ExcelReader(self.mock_file_path)
            reader.workbook = None # Ensure workbook is not loaded

            # Act & Assert
            with self.assertRaises(ExcelReadError) as cm:
                reader.get_sheet_names()
            self.assertIn("Workbook not loaded", str(cm.exception))

    def test_get_sheet_active(self):
        print(f"--- Running: {self.__class__.__name__}.test_get_sheet_active ---")
        """Test get_sheet returns the active sheet when no name is provided."""
        # Arrange
        with patch('core.reader.Path.exists', return_value=True):
            reader = ExcelReader(self.mock_file_path)
            mock_workbook = MagicMock(spec=Workbook)
            mock_active_sheet = MagicMock(spec=Worksheet, title="ActiveSheet")
            reader.workbook = mock_workbook
            reader.active_sheet = mock_active_sheet

            # Act
            sheet = reader.get_sheet() # No name provided

            # Assert
            self.assertEqual(sheet, mock_active_sheet)

    def test_get_sheet_by_name_success(self):
        print(f"--- Running: {self.__class__.__name__}.test_get_sheet_by_name_success ---")
        """Test get_sheet returns the correct sheet by name."""
        # Arrange
        with patch('core.reader.Path.exists', return_value=True):
            reader = ExcelReader(self.mock_file_path)
            mock_workbook = MagicMock(spec=Workbook)
            mock_sheet1 = MagicMock(spec=Worksheet, title="Sheet1")
            mock_sheet2 = MagicMock(spec=Worksheet, title="Sheet2")
            mock_workbook.sheetnames = ["Sheet1", "Sheet2"]
            # Mock dictionary access for workbook['SheetName']
            mock_workbook.__getitem__.side_effect = lambda key: mock_sheet2 if key == "Sheet2" else mock_sheet1
            reader.workbook = mock_workbook

            # Act
            sheet = reader.get_sheet("Sheet2")

            # Assert
            self.assertEqual(sheet, mock_sheet2)
            mock_workbook.__getitem__.assert_called_once_with("Sheet2")

    def test_get_sheet_not_found(self):
        print(f"--- Running: {self.__class__.__name__}.test_get_sheet_not_found ---")
        """Test get_sheet raises SheetNotFoundError for non-existent sheet."""
        # Arrange
        with patch('core.reader.Path.exists', return_value=True):
            reader = ExcelReader(self.mock_file_path)
            mock_workbook = MagicMock(spec=Workbook)
            mock_workbook.sheetnames = ["Sheet1", "Sheet2"]
            reader.workbook = mock_workbook

            # Act & Assert
            with self.assertRaises(SheetNotFoundError) as cm:
                reader.get_sheet("NonExistentSheet")
            self.assertIn("Sheet not found: NonExistentSheet", str(cm.exception))

    def test_get_sheet_workbook_not_loaded(self):
        print(f"--- Running: {self.__class__.__name__}.test_get_sheet_workbook_not_loaded ---")
        """Test get_sheet raises error if workbook is not loaded."""
        # Arrange
        with patch('core.reader.Path.exists', return_value=True):
            reader = ExcelReader(self.mock_file_path)
            reader.workbook = None # Ensure workbook is not loaded

            # Act & Assert
            with self.assertRaises(ExcelReadError) as cm:
                reader.get_sheet("AnySheet")
            self.assertIn("Workbook not loaded", str(cm.exception))

    # Test context manager usage
    @patch('core.reader.Path.exists', return_value=True)
    @patch.object(ExcelReader, 'load_workbook')
    @patch.object(ExcelReader, 'close')
    def test_context_manager(self, mock_close, mock_load_workbook, mock_exists):
        print(f"--- Running: {self.__class__.__name__}.test_context_manager ---")
        """Test the context manager calls load_workbook and close."""
        # Arrange
        mock_workbook = MagicMock(spec=Workbook)
        mock_load_workbook.return_value = mock_workbook
        file_path = "context_test.xlsx"

        # Act
        with ExcelReader(file_path) as reader:
            # Assertions within the context
            mock_load_workbook.assert_called_once()
            self.assertEqual(reader.workbook, mock_workbook)
            # Close should not have been called yet
            mock_close.assert_not_called()

        # Assertions after the context
        mock_close.assert_called_once()

    # TODO: Add tests for get_sheet_dimensions, get_cell_data_type


if __name__ == '__main__':
    unittest.main()