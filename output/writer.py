"""
Output writer for Excel data.
Writes formatted Excel data to various formats.
"""

import json
import logging
import os
from typing import Any, Dict, List, Optional, Union
from pathlib import Path

logger = logging.getLogger(__name__)


class WriteError(Exception):
    """Exception raised during write operations."""
    pass


class OutputWriter:
    """
    Writer for Excel data with validation support.
    
    This class provides methods for writing validated Excel data
    to various formats with error handling and validation.
    """
    
    def __init__(self):
        """Initialize the output writer."""
        pass
    
    def _ensure_directory_exists(self, file_path: str) -> None:
        """
        Ensure the directory for the output file exists.
        
        Args:
            file_path: Path to the output file
        """
        directory = os.path.dirname(file_path)
        if directory and not os.path.exists(directory):
            os.makedirs(directory, exist_ok=True)
    
    def write_json(self, data: Dict[str, Any], output_file: str, indent: int = 2) -> None:
        """
        Write data to a JSON file.
        
        Args:
            data: Data to write
            output_file: Path to the output file
            indent: Number of spaces for indentation
            
        Raises:
            WriteError: If writing fails
        """
        try:
            # Ensure directory exists
            self._ensure_directory_exists(output_file)
            
            # Write to file
            with open(output_file, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=indent)
            
            logger.info(f"Successfully wrote JSON data to {output_file}")
        
        except Exception as e:
            logger.error(f"Error writing JSON data: {e}")
            raise WriteError(f"Failed to write JSON data: {str(e)}") from e
    
    def write_csv(self, data: List[Dict[str, Any]], output_file: str, 
                  delimiter: str = ',', include_header: bool = True) -> None:
        """
        Write data to a CSV file.
        
        Args:
            data: List of dictionaries to write
            output_file: Path to the output file
            delimiter: Delimiter to use
            include_header: Whether to include a header row
            
        Raises:
            WriteError: If writing fails
        """
        try:
            import csv
            
            # Ensure directory exists
            self._ensure_directory_exists(output_file)
            
            # Get all unique keys
            keys = set()
            for row in data:
                keys.update(row.keys())
            keys = sorted(keys)
            
            # Write to file
            with open(output_file, 'w', newline='', encoding='utf-8') as f:
                writer = csv.DictWriter(f, fieldnames=keys, delimiter=delimiter)
                
                # Write header
                if include_header:
                    writer.writeheader()
                
                # Write rows
                writer.writerows(data)
            
            logger.info(f"Successfully wrote CSV data to {output_file}")
        
        except Exception as e:
            logger.error(f"Error writing CSV data: {e}")
            raise WriteError(f"Failed to write CSV data: {str(e)}") from e
    
    def write_excel(self, data: Dict[str, List[Dict[str, Any]]], output_file: str) -> None:
        """
        Write data to an Excel file.
        
        Args:
            data: Dictionary mapping sheet names to lists of row dictionaries
            output_file: Path to the output file
            
        Raises:
            WriteError: If writing fails
        """
        try:
            import pandas as pd
            from openpyxl import Workbook
            from openpyxl.utils.dataframe import dataframe_to_rows
            
            # Ensure directory exists
            self._ensure_directory_exists(output_file)
            
            # Create workbook
            wb = Workbook()
            
            # Remove default sheet
            if "Sheet" in wb.sheetnames:
                wb.remove(wb["Sheet"])
            
            # Process each sheet
            for sheet_name, rows in data.items():
                # Create DataFrame
                df = pd.DataFrame(rows)
                
                # Create sheet
                ws = wb.create_sheet(sheet_name)
                
                # Write data
                for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
                    for c_idx, value in enumerate(row, 1):
                        ws.cell(row=r_idx, column=c_idx, value=value)
            
            # Save workbook
            wb.save(output_file)
            
            logger.info(f"Successfully wrote Excel data to {output_file}")
        
        except Exception as e:
            logger.error(f"Error writing Excel data: {e}")
            raise WriteError(f"Failed to write Excel data: {str(e)}") from e